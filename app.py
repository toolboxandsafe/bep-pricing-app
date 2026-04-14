"""
BEP Pricing Calculator - Streamlit Web App
Tool Box & Safe Moving - Vending Machine Move Pricing
Replicates Email-to-Trello Workflow with Google Maps Integration
"""

import streamlit as st
import json
import math
import re
import subprocess
import tempfile
import os
import imaplib
import email
from email.header import decode_header
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
from fpdf import FPDF
import requests
import hashlib
from PyPDF2 import PdfReader, PdfWriter
from supabase import create_client

# Page config
st.set_page_config(
    page_title="BEP Pricing Calculator",
    page_icon="🚚",
    layout="wide"
)

# =============================================================================
# AUTHENTICATION
# =============================================================================
def check_password():
    """Returns True if the user has entered a correct password."""
    
    # Get credentials from environment variables
    correct_username = os.environ.get("APP_USERNAME", "admin")
    correct_password = os.environ.get("APP_PASSWORD", "bep2026")
    
    def login_form():
        """Show login form"""
        st.markdown("## 🔐 Login Required")
        st.markdown("This app is password protected.")
        
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", use_container_width=True)
            
            if submitted:
                if username == correct_username and password == correct_password:
                    st.session_state["authenticated"] = True
                    st.session_state["username"] = username
                    st.rerun()
                else:
                    st.error("❌ Invalid username or password")
        return False
    
    # Check if already authenticated
    if st.session_state.get("authenticated", False):
        return True
    
    # Show login form
    login_form()
    return False

# Require authentication before showing app
if not check_password():
    st.stop()

# =============================================================================
# MAIN APP (only runs if authenticated)
# =============================================================================

# Constants
HQ_ADDRESS = "Gilbert, AZ 85295"
HOURLY_RATE = 170

# Learning Data File
LEARNING_DATA_FILE = "learning_data.json"

# Supabase client (for route cache)
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
try:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None
except Exception as _e:
    supabase = None
    st.warning(f"Supabase client init failed: {_e}")

# Location adjustments from BEP Pricing Rulebook (413 jobs analyzed)
LOCATION_ADJUSTMENTS = {
    "far_scottsdale": {"keywords": ["north scottsdale", "n scottsdale", "n. scottsdale", "mvd north"], "adjustment": 50},
    "far_phoenix": {"keywords": ["north phoenix", "n phoenix", "anthem", "cave creek"], "adjustment": 40},
    "west_valley": {"keywords": ["peoria", "glendale", "surprise", "goodyear", "buckeye"], "adjustment": 35},
    "tucson_delivery": {"keywords": ["tucson"], "min_price": 850},
    "prison": {"keywords": ["aspc", "prison", "lewis", "perryville", "florence", "cimarron"], "adjustment": 100, "min_price": 900},
    "border": {"keywords": ["douglas", "lpoe", "san luis", "yuma border"], "adjustment": 500, "min_price": 1500},
    "rest_area": {"keywords": ["rest area", "rest stop", "sunset point"], "adjustment": 80},
}

# =============================================================================
# LEARNING DATA FUNCTIONS
# =============================================================================

def load_learning_data():
    """Load learning data from JSON file"""
    try:
        if os.path.exists(LEARNING_DATA_FILE):
            with open(LEARNING_DATA_FILE, 'r') as f:
                return json.load(f)
    except Exception as e:
        st.warning(f"Could not load learning data: {e}")
    
    # Return default structure
    return {
        "version": "1.0",
        "created": datetime.now().isoformat(),
        "quotes": [],
        "location_stats": {},
        "total_quotes": 0,
        "total_adjustments": 0,
        "avg_adjustment": 0
    }

def save_learning_data(data):
    """Save learning data to JSON file"""
    try:
        data["updated"] = datetime.now().isoformat()
        with open(LEARNING_DATA_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Could not save learning data: {e}")
        return False

def log_quote_feedback(original_quote, final_price, card_name, locations, comments=""):
    """Log feedback when quote differs from final price"""
    data = load_learning_data()
    
    diff = final_price - original_quote
    
    # Create quote record
    quote_record = {
        "timestamp": datetime.now().isoformat(),
        "card_name": card_name,
        "original_quote": original_quote,
        "final_price": final_price,
        "adjustment": diff,
        "adjustment_pct": round((diff / original_quote) * 100, 1) if original_quote > 0 else 0,
        "locations": locations,
        "comments": comments
    }
    
    data["quotes"].append(quote_record)
    data["total_quotes"] += 1
    
    if diff != 0:
        data["total_adjustments"] += 1
    
    # Update location stats
    for loc in locations:
        loc_lower = loc.lower()
        for loc_type, loc_info in LOCATION_ADJUSTMENTS.items():
            for keyword in loc_info.get("keywords", []):
                if keyword in loc_lower:
                    if loc_type not in data["location_stats"]:
                        data["location_stats"][loc_type] = {"count": 0, "total_diff": 0, "avg_diff": 0}
                    data["location_stats"][loc_type]["count"] += 1
                    data["location_stats"][loc_type]["total_diff"] += diff
                    data["location_stats"][loc_type]["avg_diff"] = round(
                        data["location_stats"][loc_type]["total_diff"] / data["location_stats"][loc_type]["count"]
                    )
    
    # Update average adjustment
    all_diffs = [q["adjustment"] for q in data["quotes"]]
    data["avg_adjustment"] = round(sum(all_diffs) / len(all_diffs), 2) if all_diffs else 0
    
    save_learning_data(data)
    return diff

def get_smart_adjustment(pickup_locations, delivery_locations):
    """Get smart price adjustment based on learned data and rulebook"""
    adjustment = 0
    min_price = 220  # Default minimum
    reasons = []
    
    all_locations = pickup_locations + delivery_locations
    location_text = " ".join(all_locations).lower()
    
    for loc_type, loc_info in LOCATION_ADJUSTMENTS.items():
        for keyword in loc_info.get("keywords", []):
            if keyword in location_text:
                if "adjustment" in loc_info:
                    adjustment += loc_info["adjustment"]
                    reasons.append(f"{loc_type}: +${loc_info['adjustment']}")
                if "min_price" in loc_info:
                    min_price = max(min_price, loc_info["min_price"])
                    reasons.append(f"{loc_type}: min ${loc_info['min_price']}")
                break  # Only apply once per location type
    
    # Check learned adjustments
    data = load_learning_data()
    for loc_type, stats in data.get("location_stats", {}).items():
        if stats["count"] >= 3 and abs(stats["avg_diff"]) > 20:
            # If we have enough data and consistent adjustments, apply learned adjustment
            for keyword in LOCATION_ADJUSTMENTS.get(loc_type, {}).get("keywords", []):
                if keyword in location_text:
                    learned_adj = stats["avg_diff"]
                    if learned_adj > 0:
                        adjustment += int(learned_adj * 0.5)  # Apply 50% of learned adjustment
                        reasons.append(f"learned({loc_type}): +${int(learned_adj * 0.5)}")
                    break
    
    return adjustment, min_price, reasons

def extract_price_adjustment_from_title(title):
    """
    Extract price adjustment from title patterns like:
    - '$350 change to $400'
    - '$450 Ryan said $400'
    - '$300 -change to $350'
    Returns (original_price, final_price) or (None, None)
    """
    if not title:
        return None, None
    
    # Pattern: $XXX change to $YYY or $XXX Ryan said $YYY
    match = re.search(r'\$(\d+)\s*(?:change to|-change to|Ryan said)\s*\$?(\d+)', title, re.IGNORECASE)
    if match:
        original = int(match.group(1))
        final = int(match.group(2))
        return original, final
    
    return None, None

def extract_price_adjustment_from_comments(comments):
    """
    Extract price adjustment from comments like:
    - 'change to $400'
    - 'changed to $350'
    - 'Ryan: $400'
    - 'should be $500'
    - 'adjust to $300'
    Returns final_price or None
    """
    if not comments:
        return None
    
    # Join all comments into one string
    all_comments = " ".join(comments) if isinstance(comments, list) else comments
    
    # Patterns to look for final price in comments
    patterns = [
        r'change(?:d)?\s*to\s*\$?(\d+)',           # change to $400, changed to 400
        r'adjust(?:ed)?\s*to\s*\$?(\d+)',          # adjust to $400
        r'should\s*be\s*\$?(\d+)',                  # should be $400
        r'Ryan[:\s]+\$?(\d+)',                      # Ryan: $400, Ryan $400
        r'final[:\s]+\$?(\d+)',                     # final: $400
        r'approved[:\s]+\$?(\d+)',                  # approved: $400
        r'price[:\s]+\$?(\d+)',                     # price: $400
    ]
    
    for pattern in patterns:
        match = re.search(pattern, all_comments, re.IGNORECASE)
        if match:
            return int(match.group(1))
    
    return None

def extract_original_quote_from_desc(description):
    """Extract original calculated quote from card description"""
    if not description:
        return None
    
    # Look for the marker we add
    match = re.search(r'\[CALC_QUOTE:(\d+)\]', description)
    if match:
        return int(match.group(1))
    
    # Fallback: look for quote in description
    match = re.search(r'###\s*💰\s*QUOTE:\s*\$?([\d,]+)', description)
    if match:
        return int(match.group(1).replace(',', ''))
    
    return None

def extract_locations_from_desc(description):
    """Extract pickup and delivery locations from card description"""
    locations = []
    if not description:
        return locations
    
    # Find pickup locations
    pickups = re.findall(r'Pickup:\s*(.+?)(?:\n|$)', description)
    locations.extend(pickups)
    
    # Find delivery locations  
    deliveries = re.findall(r'Delivery:\s*(.+?)(?:\n|$)', description)
    locations.extend(deliveries)
    
    return locations

def get_card_comments(card_id, api_key, api_token):
    """Get comments from a Trello card"""
    url = f"https://api.trello.com/1/cards/{card_id}/actions"
    params = {
        'key': api_key,
        'token': api_token,
        'filter': 'commentCard'
    }
    try:
        response = requests.get(url, params=params)
        if response.status_code == 200:
            actions = response.json()
            comments = []
            for action in actions:
                if action.get('type') == 'commentCard':
                    comment_text = action.get('data', {}).get('text', '')
                    comments.append(comment_text)
            return comments
    except Exception as e:
        pass
    return []

def post_analysis_request(card_id, original_quote, final_price, ryan_comment, api_key, api_token):
    """Post a comment requesting AI analysis from Grant"""
    diff = final_price - original_quote
    diff_pct = round((diff / original_quote) * 100, 1) if original_quote > 0 else 0
    
    comment = f"""🤖 **@grantworks2026 - AI Analysis Requested**

📊 **Price Adjustment Detected:**
- Original Quote: ${original_quote}
- Final Price: ${final_price}
- Adjustment: ${diff:+d} ({diff_pct:+.1f}%)

💬 **Ryan's Feedback:**
{ryan_comment if ryan_comment else "(No comment provided)"}

---
*Please analyze the factors contributing to this adjustment and update the learning database.*
"""
    
    url = f"https://api.trello.com/1/cards/{card_id}/actions/comments"
    params = {
        'key': api_key,
        'token': api_token,
        'text': comment
    }
    
    try:
        response = requests.post(url, params=params)
        return response.status_code == 200
    except Exception as e:
        return False

# Gmail IMAP settings
GMAIL_USER = os.environ.get("GMAIL_USER", "grantworks2026@gmail.com")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")

# =============================================================================
# GMAIL IMAP FUNCTIONS
# =============================================================================

def connect_to_gmail():
    """Connect to Gmail via IMAP"""
    if not GMAIL_APP_PASSWORD:
        return None
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        return mail
    except Exception as e:
        st.error(f"Gmail connection failed: {e}")
        return None

def get_recent_emails_with_excel(mail, limit=20):
    """Get recent emails that have Excel attachments"""
    emails = []
    try:
        mail.select("INBOX")
        
        # Search for recent emails
        _, message_numbers = mail.search(None, "ALL")
        message_list = message_numbers[0].split()
        
        # Get last 20 emails only (faster, less API calls)
        recent_messages = message_list[-20:] if len(message_list) > 20 else message_list
        recent_messages.reverse()  # Most recent first
        
        for num in recent_messages:
            if len(emails) >= limit:
                break
                
            _, msg_data = mail.fetch(num, "(RFC822)")
            email_body = msg_data[0][1]
            msg = email.message_from_bytes(email_body)
            
            # Check for Excel attachment
            has_excel = False
            for part in msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue
                filename = part.get_filename()
                if filename and (filename.endswith('.xlsx') or filename.endswith('.xls')):
                    has_excel = True
                    break
            
            if has_excel:
                # Decode subject
                subject = msg["Subject"]
                if subject:
                    decoded = decode_header(subject)
                    subject = ""
                    for part, encoding in decoded:
                        if isinstance(part, bytes):
                            subject += part.decode(encoding or 'utf-8', errors='ignore')
                        else:
                            subject += part
                
                # Get date
                date_str = msg["Date"]
                
                # Get sender
                sender = msg["From"]
                
                emails.append({
                    "id": num.decode(),
                    "subject": subject or "(No Subject)",
                    "from": sender,
                    "date": date_str,
                    "message": msg
                })
        
        return emails
    except Exception as e:
        st.error(f"Error fetching emails: {e}")
        return []

def get_excel_from_email(msg):
    """Extract Excel attachment from email message (returns first one - legacy)"""
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename = part.get_filename()
        if filename and (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return {
                "filename": filename,
                "data": part.get_payload(decode=True)
            }
    return None

def get_all_excels_from_email(msg):
    """Extract ALL Excel attachments from email message"""
    excels = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename = part.get_filename()
        if filename and (filename.endswith('.xlsx') or filename.endswith('.xls')):
            excels.append({
                "filename": filename,
                "data": part.get_payload(decode=True)
            })
    return excels

def identify_excel_type(filename, excel_bytes):
    """
    Identify if Excel is Move Request (MR) or Work Order (WO)
    Returns: 'MR', 'WO', or 'UNKNOWN'
    """
    filename_upper = filename.upper()
    
    # Check filename patterns first
    if any(x in filename_upper for x in ['WORK ORDER', 'WORKORDER', 'WO-', 'WO_', ' WO ']):
        return 'WO'
    if any(x in filename_upper for x in ['MOVE REQUEST', 'MOVEREQUEST', 'MR-', 'MR_', ' MR ', '1VR-', '1SD-']):
        return 'MR'
    
    # Check sheet names
    try:
        xl = pd.ExcelFile(BytesIO(excel_bytes))
        sheet_names_upper = [s.upper() for s in xl.sheet_names]
        
        # Work Orders typically have "WORK ORDER" sheet
        if any('WORK ORDER' in s or 'WORKORDER' in s for s in sheet_names_upper):
            return 'WO'
        
        # Move Requests have "REQUEST" sheet with pickup/delivery
        if 'REQUEST' in sheet_names_upper:
            # Double check it has pickup/delivery content
            df = pd.read_excel(BytesIO(excel_bytes), sheet_name='REQUEST', header=None)
            content = df.to_string().upper()
            if 'PICK UP SITE' in content or 'DELIVERY SITE' in content:
                return 'MR'
    except:
        pass
    
    # Default: if has REQUEST sheet with addresses, it's MR
    # Otherwise assume based on content
    try:
        df = pd.read_excel(BytesIO(excel_bytes), header=None)
        content = df.to_string().upper()
        if 'PICK UP SITE' in content and 'DELIVERY SITE' in content:
            return 'MR'
        if 'CREDIT CARD' in content or 'CARD SWAP' in content or 'CARD READER' in content:
            return 'WO'
    except:
        pass
    
    return 'UNKNOWN'

def convert_workorder_to_pdf(excel_bytes, filename):
    """Convert Work Order Excel to PDF using LibreOffice"""
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_path = os.path.join(tmpdir, filename)
        with open(excel_path, 'wb') as f:
            f.write(excel_bytes)
        
        result = subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'pdf',
            '--outdir', tmpdir, excel_path
        ], capture_output=True, text=True, timeout=60)
        
        pdf_name = os.path.splitext(filename)[0] + ".pdf"
        pdf_path = os.path.join(tmpdir, pdf_name)
        
        if os.path.exists(pdf_path):
            with open(pdf_path, 'rb') as f:
                return f.read()
    return None
BUFFER_THRESHOLD_MILES = 35
BUFFER_MINUTES = 20
JOB_TIME_PER_MACHINE = 30  # minutes
GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "")

# =============================================================================
# EXCEL PARSING - BEP REQUEST TAB (IMPROVED)
# =============================================================================

def _looks_like_full_address(text):
    """True if text contains a street number + street type (a real address)."""
    if not text or len(text) < 10:
        return False
    return bool(re.search(
        r'\d+\s+[A-Z0-9\s\.\'-]+?(?:ST|AVE|BLVD|RD|DR|LN|WAY|PKWY|HWY|STREET|AVENUE|DRIVE|ROAD|LANE|BOULEVARD|HIGHWAY|PARKWAY|CIR|CIRCLE|CT|COURT|PL|PLACE)\b',
        text.upper()
    ))

def prebuild_address_lookup(rows):
    """
    Pre-pass over all rows: find every cell that looks like a full address and
    store lookups by (a) the address's own name-prefix (text before the first
    street number), and (b) the hardcoded keyword list. This lets the main
    parser resolve name-only cells regardless of whether the full version
    appears before or after them in the sheet.
    """
    known = {}
    for row in rows:
        for cell in row:
            if not cell or not _looks_like_full_address(cell):
                continue
            store_address_keywords(cell, known)
            # Extract name-prefix: text before the first street number
            m = re.match(r'^([^\d]+?)\s+\d', cell)
            if m:
                prefix = m.group(1).strip(' -,:').upper()
                if len(prefix) >= 3 and prefix not in known:
                    known[prefix] = cell
            # Extract name-suffix: text AFTER the street suffix word, for cases
            # where the facility name comes after the address
            s = re.search(
                r'(?:STREET|ST|AVENUE|AVE|BLVD|BOULEVARD|ROAD|RD|DRIVE|DR|LANE|LN|WAY|PARKWAY|PKWY|HIGHWAY|HWY|COURT|CT|CIRCLE|CIR|PLACE|PL)\b[\s,]+([A-Za-z][A-Za-z\s]{2,40}?)(?:\s+\d|\s*$|,)',
                cell, re.IGNORECASE
            )
            if s:
                suffix = s.group(1).strip(' -,:').upper()
                # Skip if it's obviously a city/state/noise
                if (len(suffix) >= 3 and suffix not in known
                        and not re.match(r'^(PHOENIX|TUCSON|MESA|CHANDLER|GILBERT|TEMPE|AZ|ARIZONA)$', suffix)
                        and suffix.upper() not in [w.upper() for w in NOISE_TAIL_WORDS]):
                    known[suffix] = cell
    return known

def store_address_keywords(full_address, known_addresses):
    """Extract keywords from a full address and store for later lookup"""
    if not full_address or len(full_address) < 15:
        return

    addr_upper = full_address.upper()
    
    # Extract potential short names (facility names, building names)
    # Common patterns: "Bevell", "Maximus", "DES", "ADES", pool names, etc.
    keywords_to_extract = [
        'BEVELL', 'MAXIMUS', 'ADES', 'DES', 'DCS', 'MVD', 'DMV',
        'POOL', 'POLICE', 'COURT', 'CENTER', 'CIVIC', 'CAMPUS',
        'EASTLAKE', 'WESTLAKE', 'MADISON', 'CACTUS', 'KIWANIS',
        'PEORIA', 'GLENDALE', 'TEMPE', 'MESA', 'GILBERT', 'CHANDLER',
        'SCOTTSDALE', 'PHOENIX', 'TUCSON', 'YUMA', 'FLAGSTAFF'
    ]
    
    for keyword in keywords_to_extract:
        if keyword in addr_upper:
            # Store mapping: keyword -> full address
            if keyword not in known_addresses:
                known_addresses[keyword] = full_address
    
    # Also store any capitalized words that might be facility names
    words = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\b', full_address)
    for word in words:
        if len(word) > 4 and word.upper() not in ['STREET', 'AVENUE', 'DRIVE', 'ROAD', 'SUITE', 'FLOOR']:
            known_addresses[word.upper()] = full_address

def resolve_short_address(addr_text, known_addresses):
    """
    If addr_text is a short name (like 'Bevell'), try to resolve it 
    to a full address we've seen before.
    """
    if not addr_text:
        return addr_text
    
    addr_text = addr_text.strip()
    
    # If it looks like a full address (has numbers + street type), return as-is
    if re.search(r'\d+.*(?:ST|AVE|BLVD|RD|DR|LN|WAY|PKWY|HWY|STREET|AVENUE|DRIVE|ROAD)', addr_text.upper()):
        return addr_text
    
    # Short text - try to match against known addresses
    addr_upper = addr_text.upper()
    
    # Direct match
    if addr_upper in known_addresses:
        return known_addresses[addr_upper]
    
    # Partial match - check if short name is contained in any known key
    for keyword, full_addr in known_addresses.items():
        if addr_upper in keyword or keyword in addr_upper:
            return full_addr
    
    # Check if short name appears in any known full address
    for keyword, full_addr in known_addresses.items():
        if addr_upper in full_addr.upper():
            return full_addr

    # Fuzzy match against known keys to catch typos like "Conner" vs "Connor"
    from difflib import SequenceMatcher
    best_ratio = 0.0
    best_addr = None
    for keyword, full_addr in known_addresses.items():
        # Compare whole strings — fine for short facility names
        ratio = SequenceMatcher(None, addr_upper, keyword).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_addr = full_addr
    if best_ratio >= 0.85:
        return best_addr

    # No match found, return original
    return addr_text

def normalize_address(address):
    """Normalize address for deduplication"""
    if not address:
        return ""
    
    addr = str(address).upper().strip()
    
    # Remove facility prefixes
    prefixes = ['BEP ', 'MAXIMUS ', 'DES ', 'ADES ', 'DCS ']
    for prefix in prefixes:
        if addr.startswith(prefix):
            addr = addr[len(prefix):]
    
    # Extract street address pattern
    street_match = re.search(r'(\d+\s+[A-Z0-9\s\.]+(?:STREET|ST|AVENUE|AVE|BLVD|ROAD|RD|DRIVE|DR|LANE|LN|WAY|PKWY|HWY|CAMELBACK|WASHINGTON|JEFFERSON|VAN BUREN|INDIAN SCHOOL)[\.]*)', addr)
    if street_match:
        addr = street_match.group(1)
    
    # Standardize
    addr = addr.replace(' STREET', ' ST').replace(' AVENUE', ' AVE')
    addr = addr.replace(' BOULEVARD', ' BLVD').replace(' ROAD', ' RD')
    addr = addr.replace('.', '').replace(',', '').replace('#', ' ')
    addr = re.sub(r'\s+', ' ', addr).strip()
    
    return addr

def parse_bep_excel_v2(uploaded_file):
    """
    Parse BEP Move Request Excel - Extract machines with pickup/delivery pairs
    IMPORTANT: Stop parsing machines when we hit "Other Notes" section
    """
    try:
        # Read REQUEST tab
        try:
            df = pd.read_excel(uploaded_file, sheet_name='REQUEST', header=None)
        except:
            df = pd.read_excel(uploaded_file, header=None)
        
        result = {
            "success": True,
            "requester": None,
            "mr_number": None,
            "machines": [],
            "other_notes": None,
            "move_date": None,
            "contacts": [],
            "raw_data": []
        }
        
        # Convert to list of rows
        rows = []
        for idx, row in df.iterrows():
            row_data = [str(c).strip() if pd.notna(c) else "" for c in row]
            rows.append(row_data)
            result["raw_data"].append(" | ".join([c for c in row_data if c]))
        
        # FIRST PASS: Find key section boundaries
        other_notes_row = None
        machine_section_start = None
        
        for row_idx, row in enumerate(rows):
            row_text = " ".join(row).upper()
            
            # Find where "Comment" section starts - this is END of machine data
            if "COMMENT" in row_text:
                other_notes_row = row_idx
                # Capture ALL content below Comments as other_notes
                notes_lines = []
                for check_row in rows[row_idx:]:
                    for cell in check_row:
                        if cell and len(cell) > 3 and 'NAN' not in cell.upper():
                            # Skip the "Comments:" label itself
                            if cell.upper().strip() not in ['COMMENT', 'COMMENTS', 'COMMENTS:']:
                                notes_lines.append(cell.strip())
                result["other_notes"] = " | ".join(notes_lines) if notes_lines else None
                break
            
            # Also check for "Other Notes" as alternate boundary
            if "OTHER NOTE" in row_text or "SPECIAL INSTRUCTION" in row_text:
                if other_notes_row is None:
                    other_notes_row = row_idx
            
            # Find where machine data starts (first numbered item or "PICK UP" header)
            if machine_section_start is None:
                if "ITEMS TO" in row_text or "ITEM TO" in row_text:
                    machine_section_start = row_idx
        
        # Find MR number
        for row in rows:
            for cell in row:
                mr_match = re.search(r'(1\w{2}-\d{2}/\d{2})', cell)
                if mr_match:
                    result["mr_number"] = mr_match.group(1)
                    break
        
        # Find requester - try multiple strategies
        # Strategy 1: Look for "Requester Name" label and grab adjacent cell
        for row_idx, row in enumerate(rows):
            row_text = " ".join(row).upper()
            if "REQUESTER" in row_text and "NAME" in row_text:
                # Check row above for the actual name (sometimes name is above label)
                if row_idx > 0:
                    for cell in rows[row_idx - 1]:
                        if cell and len(cell) > 2 and 'NAN' not in cell.upper():
                            if not any(x in cell.upper() for x in ['REQUESTER', 'DATE', 'SIGNATURE', 'FACILITY']):
                                result["requester"] = cell
                                break
                # Also check same row for name in adjacent cell
                if not result["requester"]:
                    for col_idx, cell in enumerate(row):
                        if "REQUESTER" in cell.upper():
                            # Get next non-empty cell
                            for next_cell in row[col_idx+1:]:
                                if next_cell and len(next_cell) > 2 and 'NAN' not in next_cell.upper():
                                    result["requester"] = next_cell
                                    break
                            break
                break
        
        # Strategy 2: Fallback to row 47 (index 46)
        if not result["requester"] and len(rows) > 46:
            for cell in rows[46][:4]:
                if cell and len(cell) > 2 and not any(x in cell.upper() for x in ['REQUESTER', 'NAME', 'DATE', 'SIGNATURE', 'NAN']):
                    result["requester"] = cell
                    break
        
        # SECOND PASS: Extract machines ONLY up to "Other Notes" section
        end_row = other_notes_row if other_notes_row else len(rows)
        start_row = machine_section_start if machine_section_start else 0
        
        current_machine = None
        current_section = None
        
        # Track all full addresses we've seen (for resolving short names).
        # Pre-pass the whole sheet so forward references (short name appearing
        # before its full address) resolve correctly too.
        known_addresses = prebuild_address_lookup(rows[:end_row])
        
        for row_idx in range(start_row, end_row):
            row = rows[row_idx]
            row_text = " ".join(row).upper()
            
            # Skip if we've hit other notes
            if "OTHER NOTE" in row_text:
                break
            
            # Check for machine number (must be in specific column position, not random numbers)
            # Machine numbers are typically in column A or B (index 0 or 1)
            for col_idx in range(min(2, len(row))):
                cell = row[col_idx].strip()
                # Must be standalone digit 1-10, not part of address or date
                if cell in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']:
                    # Verify this row also has pickup/delivery context nearby
                    if any(kw in row_text for kw in ['PICK', 'DELIVER', 'SITE', 'ITEM', 'MACHINE', 'VEND', 'COMBO', 'KIOSK']):
                        if current_machine and (current_machine.get("pickup") or current_machine.get("delivery")):
                            result["machines"].append(current_machine)
                        current_machine = {
                            "number": int(cell),
                            "pickup": None,
                            "pickup_address": None,
                            "delivery": None,
                            "delivery_address": None,
                            "type": None
                        }
                        break
            
            # Extract data based on EXACT row labels
            # "Pick up site:" row contains pickup address
            # "Delivery site:" row contains delivery address  
            # "Items to be moved" row contains machine type (NOT an address)
            
            if current_machine:
                for col_idx, cell in enumerate(row):
                    cell_upper = cell.upper() if cell else ""
                    
                    # PICKUP ADDRESS: Look for "Pick up site" label, grab address from same row
                    if "PICK UP SITE" in cell_upper or "PICKUP SITE" in cell_upper:
                        # Address should be in next cell(s) on same row
                        for addr_cell in row[col_idx+1:]:
                            if addr_cell and len(addr_cell) > 3 and 'NAN' not in addr_cell.upper():
                                if not current_machine["pickup"]:
                                    resolved_addr = resolve_short_address(addr_cell, known_addresses)
                                    current_machine["pickup"] = resolved_addr
                                    current_machine["pickup_address"] = normalize_address(resolved_addr)
                                    # Store full addresses for future reference
                                    if len(addr_cell) > 20:  # Likely a full address
                                        store_address_keywords(addr_cell, known_addresses)
                                break
                    
                    # DELIVERY ADDRESS: Look for "Delivery site" label, grab address from same row
                    elif "DELIVERY SITE" in cell_upper or "DELIVER SITE" in cell_upper or "DELIVER TO" in cell_upper:
                        # Address should be in next cell(s) on same row
                        for addr_cell in row[col_idx+1:]:
                            if addr_cell and len(addr_cell) > 3 and 'NAN' not in addr_cell.upper():
                                if not current_machine["delivery"]:
                                    resolved_addr = resolve_short_address(addr_cell, known_addresses)
                                    current_machine["delivery"] = resolved_addr
                                    current_machine["delivery_address"] = normalize_address(resolved_addr)
                                    # Store full addresses for future reference
                                    if len(addr_cell) > 20:  # Likely a full address
                                        store_address_keywords(addr_cell, known_addresses)
                                break
                    
                    # MACHINE TYPE: Look for "Items to be moved" label, grab type from same row
                    elif "ITEM" in cell_upper and "MOVE" in cell_upper:
                        # Machine type should be in next cell(s) on same row
                        for type_cell in row[col_idx+1:]:
                            if type_cell and len(type_cell) > 2 and 'NAN' not in type_cell.upper():
                                if not current_machine["type"]:
                                    current_machine["type"] = type_cell
                                break
        
        # Add last machine if valid
        if current_machine and (current_machine.get("pickup") or current_machine.get("delivery")):
            result["machines"].append(current_machine)
        
        # Fallback: If no numbered machines found, try alternate method
        if not result["machines"]:
            result["machines"] = extract_machines_alternate(rows[:end_row])
        
        # Deduplicate
        result["unique_pickups"] = list(set([m["pickup_address"] for m in result["machines"] if m.get("pickup_address")]))
        result["unique_deliveries"] = list(set([m["delivery_address"] for m in result["machines"] if m.get("delivery_address")]))
        
        return result
        
    except Exception as e:
        return {"success": False, "error": str(e)}

def extract_machines_alternate(rows):
    """Alternate extraction when numbered sections aren't found"""
    machines = []
    pickups = []
    deliveries = []
    current_section = None
    
    for row in rows:
        row_text = " ".join(row).upper()
        
        if "PICK UP" in row_text or "PICKUP" in row_text:
            current_section = "pickup"
            continue
        elif "DELIVER" in row_text or "DELIVERY" in row_text:
            current_section = "delivery"
            continue
        
        for cell in row:
            if not cell or cell.upper() == 'NAN':
                continue
            cell_upper = cell.upper()
            
            is_address = any(ind in cell_upper for ind in [
                'AVE', 'STREET', 'ST ', 'BLVD', 'RD ', 'DRIVE', 'DR ', 'PHOENIX', 
                'PHX', 'TUCSON', 'MAXIMUS', 'DES ', 'BEP '
            ]) or re.search(r'\d{4,5}\s+[A-Z]', cell_upper)
            
            if is_address:
                if current_section == "pickup":
                    pickups.append(cell)
                elif current_section == "delivery":
                    deliveries.append(cell)
    
    # Pair pickups with deliveries
    num_machines = max(len(pickups), len(deliveries), 1)
    for i in range(num_machines):
        machines.append({
            "number": i + 1,
            "pickup": pickups[i] if i < len(pickups) else None,
            "pickup_address": normalize_address(pickups[i]) if i < len(pickups) else None,
            "delivery": deliveries[i] if i < len(deliveries) else None,
            "delivery_address": normalize_address(deliveries[i]) if i < len(deliveries) else None,
            "type": None
        })
    
    return machines

# =============================================================================
# GOOGLE MAPS DISTANCE MATRIX API
# =============================================================================

def get_distance_matrix(origins, destinations):
    """Call Google Maps Distance Matrix API"""
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    
    # Simple params without departure_time (avoids billing/permission issues)
    params = {
        "origins": "|".join(origins),
        "destinations": "|".join(destinations),
        "key": GOOGLE_MAPS_API_KEY,
        "units": "imperial"
    }
    
    try:
        response = requests.get(url, params=params)
        data = response.json()
        
        if data["status"] == "OK":
            return data
        else:
            st.error(f"Google Maps API error: {data.get('status')}")
            return None
    except Exception as e:
        st.error(f"API request failed: {e}")
        return None

# =============================================================================
# ROUTE CACHE (Supabase)
# =============================================================================

def _cache_key(origin, destination):
    """Stable MD5 cache key from origin|destination."""
    raw = f"{origin.strip().lower()}|{destination.strip().lower()}"
    return hashlib.md5(raw.encode()).hexdigest()

def _get_cached_route(origin, destination):
    """Look up a cached route leg from Supabase. Returns dict or None."""
    if not supabase:
        return None
    try:
        key = _cache_key(origin, destination)
        result = supabase.table("route_cache").select("distance_miles,duration_minutes").eq("cache_key", key).limit(1).execute()
        if result.data:
            return result.data[0]
    except Exception as e:
        st.warning(f"route_cache read failed: {type(e).__name__}: {e}")
    return None

def _save_cached_route(origin, destination, distance_miles, duration_minutes):
    """Store a route leg in Supabase cache."""
    if not supabase:
        return
    try:
        supabase.table("route_cache").upsert({
            "cache_key": _cache_key(origin, destination),
            "origin": origin,
            "destination": destination,
            "distance_miles": float(distance_miles),
            "duration_minutes": float(duration_minutes),
        }).execute()
    except Exception as e:
        st.warning(f"route_cache write failed: {type(e).__name__}: {e}")

AZ_CITIES = [
    'phoenix', 'tucson', 'mesa', 'chandler', 'gilbert', 'tempe', 'scottsdale',
    'glendale', 'peoria', 'yuma', 'flagstaff', 'prescott', 'surprise', 'avondale',
    'goodyear', 'buckeye', 'casa grande', 'sierra vista', 'maricopa', 'oro valley',
    'queen creek', 'kingman', 'marana', 'el mirage', 'san luis', 'sahuarita',
    'fountain hills', 'nogales', 'douglas', 'eloy', 'payson', 'show low',
    'cottonwood', 'camp verde', 'bullhead city', 'lake havasu city', 'apache junction',
]
NOISE_TAIL_WORDS = [
    'floor', 'fl', 'suite', 'ste', 'lobby', 'room', 'rm', 'BR', 'bldg', 'building',
    'unit', 'apt', 'dept', 'department', 'basement', 'bsmt', 'mezzanine',
    'office', 'wing', 'pod', 'cube', 'dock', 'warehouse',
]

def _detect_az_city(text):
    """Return the first AZ city mentioned in text, or None. Also maps PHX → Phoenix."""
    if re.search(r'\bPHX\b', text, re.IGNORECASE):
        return 'Phoenix'
    t = text.lower()
    for city in AZ_CITIES:
        if re.search(r'\b' + re.escape(city) + r'\b', t):
            return city.title()
    return None

def clean_address_for_geocoding(raw):
    """
    Strip noise from messy BEP address strings so Google Maps can geocode them.

    BEP Excel files interleave floor/room/facility-name text with the actual
    street address, e.g.:
      "MCAO 4th floor BR 225 W Madison Street PHX"
      "DES Clarendon Avenue 4000 N central 19th floor"
    This extracts the usable street portion (from the first plausible street
    number onward), strips floor/room/suite/building noise, normalizes PHX →
    Phoenix AZ, and appends a city/state hint if missing.
    """
    if not raw:
        return raw
    text = re.sub(r'\s+', ' ', str(raw).strip())

    # Detect any AZ city mentioned anywhere in the ORIGINAL text — we'll use
    # this as the fallback city instead of blindly defaulting to Phoenix.
    city = _detect_az_city(text) or 'Phoenix'

    # Find a plausible street number. A 5-digit zip at end-of-string (e.g.
    # "Gilbert, AZ 85295") is NOT a street number — skip it.
    num_match = None
    for m in re.finditer(r'\b(\d{3,})\b', text):
        num = m.group(1)
        # Skip 5-digit numbers that look like zip codes (near end or after state)
        if len(num) == 5 and re.search(r'\b(AZ|arizona)\b', text[:m.start()], re.IGNORECASE):
            continue
        num_match = m
        break

    if not num_match:
        # No street number → normalize PHX and return (may still fail, nothing to extract)
        return re.sub(r'\bPHX\b', 'Phoenix, AZ', text, flags=re.IGNORECASE).strip(' ,')

    tail = text[num_match.start():]

    # Strip trailing floor/room/suite/building noise
    noise_alt = '|'.join(NOISE_TAIL_WORDS)
    tail = re.sub(r'\s+\d+(?:st|nd|rd|th)\s+(?:' + noise_alt + r')\b.*$', '', tail, flags=re.IGNORECASE)
    tail = re.sub(r'\s+(?:' + noise_alt + r')\b.*$', '', tail, flags=re.IGNORECASE)
    tail = re.sub(r'\s+#\s*\w+\s*$', '', tail)  # trailing "# 102" at end only

    # Normalize PHX → Phoenix, AZ
    tail = re.sub(r'\bPHX\b', 'Phoenix, AZ', tail, flags=re.IGNORECASE)

    # Append city/state hint if nothing identifiable is present in the tail
    az_pattern = r'\b(' + '|'.join(re.escape(c) for c in AZ_CITIES) + r'|AZ|arizona)\b'
    if not re.search(az_pattern, tail, re.IGNORECASE):
        tail = tail.rstrip(' ,') + f', {city}, AZ'

    return re.sub(r'\s+', ' ', tail).strip(' ,')

def _dedupe_key(text):
    """Aggressive normalization for dedup: lowercase, strip ALL punctuation,
    collapse whitespace. So '225 W Madison St, Phoenix, AZ' and
    '225 W Madison St Phoenix AZ' hash to the same key."""
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9]+', ' ', (text or '').lower())).strip()

def clean_and_dedupe_addresses(items):
    """Clean each address for geocoding, then dedupe while preserving order."""
    cleaned = [clean_address_for_geocoding(i) for i in items if i]
    seen = {}
    for it in cleaned:
        key = _dedupe_key(it)
        if key and key not in seen:
            seen[key] = it
    return list(seen.values())

def _get_leg(origin, destination):
    """Return (distance_miles, duration_minutes) for one leg — cache-aware."""
    cached = _get_cached_route(origin, destination)
    if cached:
        return float(cached["distance_miles"]), float(cached["duration_minutes"])
    data = get_distance_matrix([origin], [destination])
    if data and data["rows"][0]["elements"][0]["status"] == "OK":
        el = data["rows"][0]["elements"][0]
        dm = el["distance"]["value"] / 1609.34
        dur = el["duration"]["value"] / 60
        _save_cached_route(origin, destination, round(dm, 1), round(dur, 1))
        return dm, dur
    return None, None

def calculate_optimal_route(machines):
    """
    Precedence-constrained TSP: visit each unique location once, minimize
    total drive time. Constraint: for each machine, its pickup location must
    be visited at or before its delivery location (same-stop is allowed).

    Returns the same dict shape as calculate_route so calculate_quote works.
    """
    # Clean each machine's pickup/delivery
    pairs = []
    for mch in machines:
        p = clean_address_for_geocoding(mch.get("pickup"))
        d = clean_address_for_geocoding(mch.get("delivery"))
        if p and d:
            pairs.append((p, d))

    if not pairs:
        return {"legs": [], "total_duration_minutes": 0, "max_distance_miles": 0, "route": [HQ_ADDRESS, HQ_ADDRESS]}

    # Build unique location index (dedupe by punctuation-insensitive key)
    locs = []          # index -> canonical address string
    key_to_idx = {}    # dedupe key -> index
    def _idx(addr):
        k = _dedupe_key(addr)
        if k not in key_to_idx:
            key_to_idx[k] = len(locs)
            locs.append(addr)
        return key_to_idx[k]

    precedence = []  # list of (pickup_idx, delivery_idx)
    for p, d in pairs:
        pi, di = _idx(p), _idx(d)
        if pi != di:
            precedence.append((pi, di))

    n = len(locs)

    # Fetch full (n+1) × (n+1) duration/distance matrix (HQ is index 0)
    nodes = [HQ_ADDRESS] + locs
    N = n + 1
    dur = [[0.0] * N for _ in range(N)]
    dist = [[0.0] * N for _ in range(N)]
    for i in range(N):
        for j in range(N):
            if i == j:
                continue
            dm, dt = _get_leg(nodes[i], nodes[j])
            if dm is None:
                dm, dt = 20.0, 30.0  # fallback estimate
            dist[i][j] = dm
            dur[i][j] = dt

    # Brute-force best permutation of loc indices satisfying precedence.
    # n! grows fast but for n<=10 it's instant; for n<=12 it's still seconds.
    from itertools import permutations
    best_perm = None
    best_dur = float("inf")

    # Precompute which indices have predecessors, to prune earlier
    preds = {i: set() for i in range(n)}
    for b, a in precedence:
        preds[a].add(b)

    for perm in permutations(range(n)):
        # Precedence check
        seen = set()
        ok = True
        for idx in perm:
            if not preds[idx].issubset(seen):
                ok = False
                break
            seen.add(idx)
        if not ok:
            continue
        # Total duration: HQ -> perm[0] -> ... -> perm[-1] -> HQ
        total = dur[0][perm[0] + 1]
        for i in range(n - 1):
            total += dur[perm[i] + 1][perm[i + 1] + 1]
        total += dur[perm[-1] + 1][0]
        if total < best_dur:
            best_dur = total
            best_perm = perm

    if best_perm is None:
        # No valid ordering — fall back to original simple routing
        return calculate_route([p for p, _ in pairs], [d for _, d in pairs])

    # Build legs list in the best order
    sequence = [0] + [i + 1 for i in best_perm] + [0]
    legs = []
    total_duration = 0.0
    max_dist = 0.0
    for i in range(len(sequence) - 1):
        a, b = sequence[i], sequence[i + 1]
        dm, dt = dist[a][b], dur[a][b]
        legs.append({
            "from": nodes[a],
            "to": nodes[b],
            "distance_miles": round(dm, 1),
            "duration_minutes": round(dt, 1),
        })
        total_duration += dt
        max_dist = max(max_dist, dm)

    return {
        "legs": legs,
        "total_duration_minutes": round(total_duration, 1),
        "max_distance_miles": round(max_dist, 1),
        "route": [nodes[i] for i in sequence],
    }

def calculate_route(pickups, deliveries):
    """
    Calculate sequential route: HQ → Pickups → Deliveries → HQ
    Returns list of legs with distance and duration
    """
    # Clean + dedupe BEFORE routing so "PV pool" and "PV Pool 17648 N 40th St"
    # collapse to one stop.
    pickups = clean_and_dedupe_addresses(pickups)
    deliveries = clean_and_dedupe_addresses(deliveries)

    # Build route
    route = [HQ_ADDRESS]
    route.extend(pickups)
    route.extend(deliveries)
    route.append(HQ_ADDRESS)

    legs = []
    total_duration_minutes = 0
    max_distance_miles = 0

    # Calculate each consecutive leg
    for i in range(len(route) - 1):
        origin = route[i]
        destination = route[i + 1]

        # Check Supabase cache first
        cached = _get_cached_route(origin, destination)
        if cached:
            distance_miles = float(cached["distance_miles"])
            duration_minutes = float(cached["duration_minutes"])
            legs.append({
                "from": origin,
                "to": destination,
                "distance_miles": round(distance_miles, 1),
                "duration_minutes": round(duration_minutes, 1),
                "cached": True,
            })
            total_duration_minutes += duration_minutes
            max_distance_miles = max(max_distance_miles, distance_miles)
            continue

        data = get_distance_matrix([origin], [destination])

        if data and data["rows"][0]["elements"][0]["status"] == "OK":
            element = data["rows"][0]["elements"][0]

            # Distance in miles
            distance_meters = element["distance"]["value"]
            distance_miles = distance_meters / 1609.34

            # Duration in minutes
            duration_seconds = element["duration"]["value"]
            duration_minutes = duration_seconds / 60

            # Save to cache for future requests
            _save_cached_route(origin, destination, round(distance_miles, 1), round(duration_minutes, 1))

            legs.append({
                "from": origin,
                "to": destination,
                "distance_miles": round(distance_miles, 1),
                "duration_minutes": round(duration_minutes, 1),
                "cached": False,
            })

            total_duration_minutes += duration_minutes
            max_distance_miles = max(max_distance_miles, distance_miles)
        else:
            st.warning(f"Could not calculate: {origin} → {destination}")
            # Estimate fallback
            legs.append({
                "from": origin,
                "to": destination,
                "distance_miles": 20,
                "duration_minutes": 30,
                "estimated": True
            })
            total_duration_minutes += 30
    
    return {
        "legs": legs,
        "total_duration_minutes": round(total_duration_minutes, 1),
        "max_distance_miles": round(max_distance_miles, 1),
        "route": route
    }

# =============================================================================
# QUOTE CALCULATION
# =============================================================================

def calculate_quote(route_data, num_machines, pickups, deliveries):
    """Calculate BEP quote following workflow rules with smart adjustments"""
    
    drive_time = route_data["total_duration_minutes"]
    max_distance = route_data["max_distance_miles"]
    
    # Job time: 30 minutes per machine
    job_time = JOB_TIME_PER_MACHINE * num_machines
    
    # Buffer: 20 minutes if max distance > 35 miles
    if max_distance > BUFFER_THRESHOLD_MILES:
        buffer_time = BUFFER_MINUTES
        no_buffer_discount = 0
    else:
        buffer_time = 0
        no_buffer_discount = 60  # -$60 for short trips
    
    # Total time
    total_minutes = drive_time + job_time + buffer_time
    total_hours = total_minutes / 60
    
    # Base price
    base_price = total_hours * HOURLY_RATE
    
    # Apply no-buffer discount
    base_price -= no_buffer_discount
    
    # Check locations for minimums
    all_locations = " ".join(pickups + deliveries).upper()
    
    is_tucson = "TUCSON" in all_locations
    is_prison = any(kw in all_locations for kw in ['ASPC', 'PRISON', 'CORRECTIONAL', 'CIMARRON', 'FLORENCE'])
    
    # Determine minimum
    if is_tucson:
        min_price = 850
    elif is_prison:
        min_price = 900
    else:
        min_price = 220
    
    # Get smart adjustments from learning system
    smart_adjustment, smart_min, adjustment_reasons = get_smart_adjustment(pickups, deliveries)
    min_price = max(min_price, smart_min)
    
    # Apply minimum
    final_price = max(base_price, min_price)
    
    # Add smart adjustment
    final_price += smart_adjustment
    
    # Round UP to nearest $25
    final_price = math.ceil(final_price / 25) * 25
    
    return {
        "drive_time": round(drive_time, 1),
        "job_time": job_time,
        "buffer_time": buffer_time,
        "no_buffer_discount": no_buffer_discount,
        "total_minutes": round(total_minutes, 1),
        "total_hours": round(total_hours, 2),
        "max_distance_miles": round(max_distance, 1),
        "base_price": round(base_price, 2),
        "min_price": min_price,
        "final_price": int(final_price),
        "is_tucson": is_tucson,
        "is_prison": is_prison,
        "smart_adjustment": smart_adjustment,
        "adjustment_reasons": adjustment_reasons,
        "formula": f"({round(drive_time)} + {job_time} + {buffer_time}) ÷ 60 × ${HOURLY_RATE} - ${no_buffer_discount} + ${smart_adjustment} (smart) = ${final_price:.0f}"
    }

# =============================================================================
# EXCEL TO PDF CONVERSION
# =============================================================================

def convert_excel_to_pdf(excel_bytes, filename="request.xlsx"):
    """Convert Excel file to PDF using LibreOffice"""
    with tempfile.TemporaryDirectory() as tmpdir:
        excel_path = os.path.join(tmpdir, filename)
        with open(excel_path, 'wb') as f:
            f.write(excel_bytes)
        
        try:
            result = subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', tmpdir, excel_path
            ], capture_output=True, text=True, timeout=60)
            
            pdf_name = os.path.splitext(filename)[0] + '.pdf'
            pdf_path = os.path.join(tmpdir, pdf_name)
            
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    return f.read()
            return None
        except Exception as e:
            st.error(f"LibreOffice conversion failed: {e}")
            return None

def remove_pdf_pages(pdf_bytes, pages_to_remove=[1]):
    """Remove specific pages from PDF (0-indexed)"""
    try:
        reader = PdfReader(BytesIO(pdf_bytes))
        writer = PdfWriter()
        
        for i, page in enumerate(reader.pages):
            if i not in pages_to_remove:
                writer.add_page(page)
        
        output = BytesIO()
        writer.write(output)
        return output.getvalue()
    except Exception as e:
        st.error(f"Error removing pages: {e}")
        return pdf_bytes

# =============================================================================
# TRELLO INTEGRATION
# =============================================================================

def create_trello_card(data, api_key, api_token, list_id):
    """Create Trello card with quote data"""
    
    # Build driving stops — prefer the optimized route sequence if present,
    # otherwise fall back to the old pickups-then-deliveries ordering.
    route_seq = data.get("route") or []
    if route_seq:
        def _short(addr):
            if not addr:
                return "?"
            if "gilbert" in addr.lower() and "85295" in addr:
                return "HQ"
            # First token that contains a digit (street number), else first word
            for tok in addr.split():
                if any(c.isdigit() for c in tok):
                    return tok[:6].upper()
            return addr.split()[0][:6].upper()
        stops = [_short(a) for a in route_seq]
        unique_stop_count = len({_dedupe_key(a) for a in route_seq if a and "gilbert" not in a.lower()})
    else:
        stops = ["HQ"]
        for p in data.get("unique_pickups", []):
            stops.append(p.split()[0][:6].upper() if p else "?")
        for d in data.get("unique_deliveries", []):
            stops.append(d.split()[0][:6].upper() if d else "?")
        stops.append("HQ")
        all_keys = {_dedupe_key(a) for a in data.get("unique_pickups", []) + data.get("unique_deliveries", []) if a}
        all_keys.discard("")
        unique_stop_count = len(all_keys)
    driving_stops = " - ".join(stops)
    
    # Use custom title if provided, otherwise build from data
    if data.get('card_title'):
        title = data.get('card_title')
    else:
        title = f"{data.get('requester', 'Unknown')} - {data.get('mr_number', 'BEP')} - ${data.get('final_price', 0)}"
    
    desc = f"""## Move Request Quote

**MR Number:** {data.get('mr_number', 'N/A')}
**Requester:** {data.get('requester', 'N/A')}
**Move Date:** {data.get('move_date', 'TBD')}
**Machines:** {data.get('num_machines', 1)}

---

### 📍 MACHINES & LOCATIONS
"""
    
    for m in data.get("machines", []):
        desc += f"\n**Machine {m.get('number', '?')}:** {m.get('type', 'Vending')}\n"
        desc += f"  - Pickup: {m.get('pickup', 'N/A')}\n"
        desc += f"  - Delivery: {m.get('delivery', 'N/A')}\n"
    
    desc += f"""
---

### 🚗 DRIVING STOPS
{driving_stops}

({unique_stop_count} unique stops)

---

### 💰 QUOTE: ${data.get('final_price', 0):,}

**Breakdown:**
- Drive Time: {data.get('drive_time', 0)} min
- Job Time: {data.get('job_time', 0)} min ({data.get('num_machines', 1)} machines × 30 min)
- Buffer: {data.get('buffer_time', 0)} min
- Max Distance Leg: {data.get('max_distance_miles', 0)} miles
- Total Hours: {data.get('total_hours', 0)}
- Rate: ${HOURLY_RATE}/hour

**Formula:** {data.get('formula', '')}

---

### 📝 OTHER NOTES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{data.get('other_notes', 'None')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

---
@luissaravia2

[CALC_QUOTE:{data.get('final_price', 0)}]
"""
    
    url = "https://api.trello.com/1/cards"
    params = {
        'key': api_key,
        'token': api_token,
        'idList': list_id,
        'name': title,
        'desc': desc,
        'pos': 'top'
    }
    
    response = requests.post(url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        st.error(f"Trello API error {response.status_code}: {response.text}")
        return None

def attach_file_to_card(card_id, file_bytes, filename, mime_type, api_key, api_token):
    """Attach any file to a Trello card"""
    url = f"https://api.trello.com/1/cards/{card_id}/attachments"
    
    params = {
        'key': api_key,
        'token': api_token,
        'name': filename
    }
    
    files = {
        'file': (filename, file_bytes, mime_type)
    }
    
    response = requests.post(url, params=params, files=files)
    if response.status_code == 200:
        return True
    else:
        st.error(f"Attachment error: {response.status_code}: {response.text}")
        return False

def attach_pdf_to_card(card_id, pdf_bytes, filename, api_key, api_token):
    """Attach PDF file to a Trello card"""
    return attach_file_to_card(card_id, pdf_bytes, filename, 'application/pdf', api_key, api_token)

def attach_excel_to_card(card_id, excel_bytes, filename, api_key, api_token):
    """Attach Excel file to a Trello card"""
    return attach_file_to_card(card_id, excel_bytes, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', api_key, api_token)

def get_card_attachments(card_id, api_key, api_token):
    """Get attachments from a Trello card"""
    url = f"https://api.trello.com/1/cards/{card_id}/attachments"
    params = {'key': api_key, 'token': api_token}
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    return []

def get_card_info(card_id, api_key, api_token):
    """Get card info including name"""
    url = f"https://api.trello.com/1/cards/{card_id}"
    params = {'key': api_key, 'token': api_token, 'fields': 'name,desc,shortUrl'}
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    return None

def download_attachment(url, api_key=None, api_token=None):
    """Download file from URL with optional Trello auth"""
    headers = {}
    
    # For Trello attachments, try with auth header
    if api_key and api_token:
        headers['Authorization'] = f'OAuth oauth_consumer_key="{api_key}", oauth_token="{api_token}"'
    
    try:
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 200:
            return response.content
        
        # Fallback: try adding auth as query params
        if api_key and api_token:
            if '?' in url:
                auth_url = f"{url}&key={api_key}&token={api_token}"
            else:
                auth_url = f"{url}?key={api_key}&token={api_token}"
            response = requests.get(auth_url, timeout=30)
            if response.status_code == 200:
                return response.content
        
        st.error(f"Download failed: {response.status_code}")
        return None
    except Exception as e:
        st.error(f"Download error: {e}")
        return None

def extract_quote_from_title(title):
    """Extract quote amount from card title like 'Name - MR# - $325'"""
    match = re.search(r'\$(\d+)', title)
    if match:
        return int(match.group(1))
    return None

def fill_worksheet_and_generate_pdf(excel_bytes, quote_amount, signature="Ryan Kearl"):
    """Fill Worksheet tab with hours and generate PDF"""
    import openpyxl
    
    # Calculate hours
    hours = round(quote_amount / HOURLY_RATE, 2)
    today_date = datetime.now().strftime("%m/%d/%Y")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Save Excel
        excel_path = os.path.join(tmpdir, "workbook.xlsx")
        with open(excel_path, 'wb') as f:
            f.write(excel_bytes)
        
        # Open and modify
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            # Find Worksheet tab (might be named differently)
            ws_names = [s for s in wb.sheetnames if 'WORKSHEET' in s.upper() or 'WORK' in s.upper()]
            if ws_names:
                ws = wb[ws_names[0]]
            else:
                # Fallback to second sheet
                ws = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]
            
            # Fill hours in row 6, columns D and I
            ws['D6'] = hours
            ws['I6'] = hours
            
            # Fill signature and date
            ws['K19'] = signature
            ws['K21'] = today_date
            
            # Save modified workbook
            modified_path = os.path.join(tmpdir, "modified.xlsx")
            wb.save(modified_path)
            wb.close()
            
            # Convert to PDF using LibreOffice
            result = subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', tmpdir, modified_path
            ], capture_output=True, text=True, timeout=60)
            
            pdf_path = os.path.join(tmpdir, "modified.pdf")
            
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    pdf_bytes = f.read()
                
                # Extract just the Worksheet page (usually page 2, index 1)
                try:
                    reader = PdfReader(BytesIO(pdf_bytes))
                    writer = PdfWriter()
                    
                    # Add only page 2 (Worksheet) if exists, else page 1
                    if len(reader.pages) > 1:
                        writer.add_page(reader.pages[1])
                    else:
                        writer.add_page(reader.pages[0])
                    
                    output = BytesIO()
                    writer.write(output)
                    return output.getvalue(), hours
                except:
                    return pdf_bytes, hours
            
            return None, hours
            
        except Exception as e:
            st.error(f"Error filling worksheet: {e}")
            return None, hours

# =============================================================================
# INVOICE GENERATION (Google Sheets + Trello)
# =============================================================================

INVOICE_SPREADSHEET_ID = os.environ.get(
    "INVOICE_SPREADSHEET_ID",
    "1UUqoUeX_3I4qzyTJt9tzdax2T7vUyVPydXWT58k1FS0",
)
INVOICE_TEMPLATE_TAB = "TEMPLATE"   # "Move Template"
INVOICE_PENDING_TAB = "Pending Payments"
INVOICE_BEP_COMPLETED_LIST = "BEP Completed"

# Templates the user can pick on the Generate INV page.
# For now only the move template is wired up.
INVOICE_JOB_TYPES = {
    "Move Template": INVOICE_TEMPLATE_TAB,
    # "Install & Setup": "INSTAL N SET UP TEMPLATE",  # future
    # "Repair": "REPAIR TEMPLATE",                     # future
}

def _get_gspread_client():
    """
    Load service account JSON from the GOOGLE_SERVICE_ACCOUNT_JSON env var
    and return an authorized gspread client. Returns (client, creds) or (None, None).
    """
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not sa_json:
        return None, None
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        info = json.loads(sa_json)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(info, scopes=scopes)
        return gspread.authorize(creds), creds
    except Exception as e:
        st.error(f"Failed to load Google service account: {type(e).__name__}: {e}")
        return None, None

def clean_card_title_for_tab(title, bep_auth=None):
    """
    Clean Trello card title for use as a sheet-tab suffix.
    - If title contains Fwd:/Fw:/Re:, keep ONLY the part AFTER the LAST such marker
      (e.g. "Robert Jeffrey - Fwd: Move request" → "Move request")
    - Strip price tokens: $N, $xxx, "$N change to $M", with commas/decimals
    - If a BEP authorization number (bep_auth) is provided and missing from
      the result, append it — auth # must ALWAYS appear in the tab name
    - Truncate to leave room for the "INV##### " prefix (~90 chars for the suffix)
    """
    if not title:
        return bep_auth or ""
    t = str(title)

    # Keep only the text after the LAST Fwd:/Fw:/Re:
    marker_matches = list(re.finditer(r'\b(?:fwd?|re)\s*:\s*', t, re.IGNORECASE))
    if marker_matches:
        t = t[marker_matches[-1].end():]

    # Remove "$N change to $M" patterns
    t = re.sub(
        r'\$?[\d,]+(?:\.\d+)?\s*change\s*to\s*\$?[\d,]+(?:\.\d+)?',
        '', t, flags=re.IGNORECASE,
    )
    # Remove "$xxx" placeholder and bare "$N"
    t = re.sub(r'\$\s*x+', '', t, flags=re.IGNORECASE)
    t = re.sub(r'\$\s*[\d,]+(?:\.\d+)?', '', t)

    # Clean up separator artifacts and whitespace
    t = re.sub(r'\s*-\s*-\s*', ' - ', t)
    t = re.sub(r'\s+', ' ', t).strip(' -,:')

    # Ensure the auth number is present in the tab name
    if bep_auth and bep_auth.upper() not in t.upper():
        t = f"{t} {bep_auth}".strip() if t else bep_auth

    return t[:80]

def extract_final_price_from_title(title):
    """
    Return the LAST dollar amount mentioned in the title as a float.
    Handles: $400, $1,200, $1,200.50, "$400 change to $500" → 500.
    Returns None if nothing looks like a real price (placeholders like $xxx ignored).
    """
    if not title:
        return None
    # Find all $N tokens, ignoring $xxx placeholders
    matches = re.findall(r'\$\s*([\d,]+(?:\.\d+)?)', str(title))
    if not matches:
        return None
    try:
        return float(matches[-1].replace(",", ""))
    except ValueError:
        return None

def _format_move_date(dt):
    """Format date like 'March 30th, 2026' — matches template style, and the
    ordinal suffix prevents Google Sheets from parsing it as a date value."""
    day = dt.day
    if 11 <= day <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return dt.strftime(f"%B {day}{suffix}, %Y")

def extract_bep_auth_from_title(title):
    """
    Extract the BEP authorization number from the card title.
    Pattern: starts with '77', ends with 'A1' (case-insensitive), alphanumeric in between.
    Example: '777037858A1'
    """
    if not title:
        return None
    m = re.search(r'\b(77\w*?[aA]1)\b', str(title))
    return m.group(1).upper() if m else None

def parse_machines_from_card_desc(desc):
    """
    Parse machines from a Trello card description. Handles multiple formats:

      Format A (old app output, markdown):
        **Machine 1:** type
          - Pickup: X
          - Delivery: Y

      Format B (current standardized):
        Machine 1: type

            Pickup: X

            Delivery: Y

    Also tolerates missing ** around "Machine N:", blank lines between rows,
    bullets like '-' or '•' before Pickup/Delivery, and varying whitespace.
    Returns list of dicts with keys: number, type, pickup, delivery.
    """
    machines = []
    if not desc:
        return machines

    # Anchor on the MACHINES & LOCATIONS section if present. Otherwise scan
    # the whole desc. Stop at the next major section header (DRIVING STOPS,
    # QUOTE, FORMULA, etc.) or end of desc.
    section = desc
    anchor = re.search(r'MACHINES\s*&\s*LOCATIONS', desc, re.IGNORECASE)
    if anchor:
        start = anchor.end()
        end_m = re.search(
            r'\n\s*(?:#+\s*)?(?:\*+\s*)?(DRIVING\s*STOPS|QUOTE|FORMULA|BREAKDOWN|NOTES?|OTHER)',
            desc[start:], re.IGNORECASE,
        )
        end = start + end_m.start() if end_m else len(desc)
        section = desc[start:end]

    # Flexible per-machine regex:
    #   **Machine N:**  or  Machine N:
    #   type on same line (may include **)
    #   one or two newlines
    #   optional bullet [-•*] and optional whitespace
    #   Pickup: ...
    #   one or two newlines
    #   optional bullet
    #   Delivery: ...
    # --- Format 1: "Machine N: type / Pick Up / Delivery" (current standard) ---
    # Handles: **Machine 1:** type, **Pick Up:**, **Delivery:**, Pickup, Pick-Up, etc.
    new_pattern = re.compile(
        r'\*{0,2}\s*Machine\s*(\d+)\s*:?\s*\*{0,2}\s*([^\n]+?)\s*\n+'
        r'\s*\*{0,2}\s*[-•]?\s*Pick\s*[-_]?\s*Up\s*(?:Site)?\s*:\s*\*{0,2}\s*([^\n]+?)\s*\n+'
        r'\s*\*{0,2}\s*[-•]?\s*Delivery\s*(?:Site)?\s*:\s*\*{0,2}\s*([^\n]+)',
        re.IGNORECASE,
    )
    for m in new_pattern.finditer(section):
        type_text = m.group(2).strip().strip("*").strip()
        pickup = m.group(3).strip().strip("*").strip()
        delivery = m.group(4).strip().strip("*").strip()
        if not pickup or not delivery or pickup.lower() in ("n/a", "none") or delivery.lower() in ("n/a", "none"):
            continue
        machines.append({
            "number": m.group(1).strip(),
            "type": type_text,
            "pickup": pickup,
            "delivery": delivery,
        })
    if machines:
        return machines

    # --- Format 2: "Items to be Moved / Pick Up Site / Delivery Site" (old) ---
    # Example:
    #   Items to be Moved: USI Alpine 3000 frozen DE3000038 S#1423227 (1 machine)
    #   Pick Up Site: ASPC Phoenix Aspen Unit, 2500 E Van Buren Street, PHX 85008
    #   Delivery Site: BEP Maximus, 3425 E Van Buren Street #102, PHX 85008
    old_pattern = re.compile(
        r'\*{0,2}\s*Items?\s+to\s+be\s+Moved\s*:\s*\*{0,2}\s*([^\n]+?)\s*\n+'
        r'\s*[-•*]?\s*Pick\s*[-]?\s*[Uu]p\s*Site\s*:\s*([^\n]+?)\s*\n+'
        r'\s*[-•*]?\s*Delivery\s*Site\s*:\s*([^\n]+)',
        re.IGNORECASE,
    )
    for i, m in enumerate(old_pattern.finditer(section), start=1):
        type_text = m.group(1).strip().strip("*").strip()
        pickup = m.group(2).strip().strip("*").strip()
        delivery = m.group(3).strip().strip("*").strip()
        if not pickup or not delivery or pickup.lower() in ("n/a", "none") or delivery.lower() in ("n/a", "none"):
            continue
        machines.append({
            "number": str(i),
            "type": type_text,
            "pickup": pickup,
            "delivery": delivery,
        })
    return machines

def extract_total_hours_from_desc(desc):
    """
    Sum Drive Time + Job Time + Buffer from the card desc and divide by 60.
    Returns a float rounded to 2 decimals, or None if nothing found.
    """
    if not desc:
        return None
    def _num(label):
        m = re.search(label + r'[:\s]*(\d+(?:\.\d+)?)\s*min', desc, re.IGNORECASE)
        return float(m.group(1)) if m else 0.0
    drive = _num(r'Drive\s*Time')
    job = _num(r'Job\s*Time')
    buf = _num(r'Buffer')
    total = drive + job + buf
    if total <= 0:
        return None
    return round(total / 60.0, 2)

def find_move_to_list_date(card_id, target_list_name, api_key, api_token):
    """
    Return the ISO date string of the most recent action where this card
    was moved INTO a list named target_list_name. Returns None if not found.
    """
    url = f"https://api.trello.com/1/cards/{card_id}/actions"
    params = {
        "key": api_key,
        "token": api_token,
        "filter": "updateCard:idList",
        "limit": 1000,
    }
    try:
        r = requests.get(url, params=params, timeout=30)
        if r.status_code != 200:
            return None
        actions = r.json()
    except Exception:
        return None
    target_lower = target_list_name.strip().lower()
    for action in actions:  # Trello returns newest first
        data = action.get("data", {})
        list_after = (data.get("listAfter") or {}).get("name", "")
        if list_after.strip().lower() == target_lower:
            return action.get("date")
    return None

def _col_last_data_row(ws, col_index=1):
    """Find the last row in a column that has a non-empty value (1-indexed)."""
    values = ws.col_values(col_index)
    for i in range(len(values), 0, -1):
        if values[i - 1] not in (None, ""):
            return i
    return 0

def get_next_invoice_number(spreadsheet):
    """Scan all tab names for INV##### pattern, return max + 1."""
    max_num = 10476  # floor: the highest currently-existing invoice at time of writing
    for ws in spreadsheet.worksheets():
        m = re.match(r'^INV(\d{4,6})\b', ws.title)
        if m:
            try:
                n = int(m.group(1))
                if n > max_num:
                    max_num = n
            except ValueError:
                pass
    return max_num + 1

def _fill_machine_block(ws, block_row, machine_number, machine):
    """
    Fill one 4-row machine block. block_row is the header row (A12, A16, ...).
    Layout per template:
      A{r}   "{n}.0"       B{r}   "Item(s) to be Moved:"   C{r}  type       I{r}  "1.0"
      B{r+1} "Pick Up Site:"                                 C{r+1} pickup
      B{r+2} "Delivery Site:"                                C{r+2} delivery
    """
    updates = [
        {"range": f"A{block_row}",    "values": [[f"{machine_number}.0"]]},
        {"range": f"B{block_row}",    "values": [["Item(s) to be Moved:"]]},
        {"range": f"C{block_row}",    "values": [[machine.get("type", "")]]},
        {"range": f"H{block_row}",    "values": [["Qty"]]},
        {"range": f"I{block_row}",    "values": [["1.0"]]},
        {"range": f"B{block_row + 1}", "values": [["Pick Up Site:"]]},
        {"range": f"C{block_row + 1}", "values": [[machine.get("pickup", "")]]},
        {"range": f"B{block_row + 2}", "values": [["Delivery Site:"]]},
        {"range": f"C{block_row + 2}", "values": [[machine.get("delivery", "")]]},
    ]
    ws.batch_update(updates, value_input_option="USER_ENTERED")

def generate_invoice_from_card(card_id, job_type_label, note_text, api_key, api_token):
    """
    End-to-end invoice generation:
      1. Fetch card (name, desc, idBoard)
      2. Parse title/desc for auth#, machines, hours
      3. Find move-to-BEP-Completed date (fallback: today)
      4. Duplicate template in the Google Sheet, place to right of TEMPLATE
      5. Fill cells, insert extra machine blocks if needed
      6. Export new tab as PDF
      7. Attach PDF to Trello card
      8. Append row to Pending Payments
    Returns dict with keys: ok, invoice_number, tab_name, tab_gid, pdf_bytes, pdf_filename,
    sheet_url, total_amount, warnings (list), error (str or None).
    """
    result = {
        "ok": False, "invoice_number": None, "tab_name": None, "tab_gid": None,
        "pdf_bytes": None, "pdf_filename": None, "sheet_url": None,
        "total_amount": None, "warnings": [], "error": None,
    }

    # --- 1. Fetch card ---
    if not api_key or not api_token:
        result["error"] = "Trello credentials missing (TRELLO_API_KEY / TRELLO_TOKEN)."
        return result
    card_info = get_card_info(card_id, api_key, api_token)
    if not card_info:
        result["error"] = f"Could not fetch Trello card {card_id}."
        return result

    # Need idBoard too — get_card_info only fetches name/desc/shortUrl
    try:
        full = requests.get(
            f"https://api.trello.com/1/cards/{card_id}",
            params={"key": api_key, "token": api_token, "fields": "name,desc,idBoard,shortUrl"},
            timeout=30,
        ).json()
    except Exception as e:
        result["error"] = f"Trello API error: {e}"
        return result

    card_name = full.get("name", "") or ""
    card_desc = full.get("desc", "") or ""
    card_short_url = full.get("shortUrl", "")

    # --- 2. Parse card data ---
    bep_auth = extract_bep_auth_from_title(card_name)
    if not bep_auth:
        result["warnings"].append(
            "Could not find BEP authorization # (77…A1) in card title. F7 will be blank."
        )
    machines = parse_machines_from_card_desc(card_desc)
    if not machines:
        # Stash the raw desc in the result so the UI can show it for debugging
        result["raw_desc"] = card_desc
        result["error"] = (
            "No machines found in card description. Cannot generate invoice. "
            "Expand the 'Raw card description' section below to see what the parser saw."
        )
        return result

    # Final price comes from the TITLE. If title has "$400 change to $500", use $500.
    # Invoice hours are computed as price / 170 (J column × $130 driver + K × $40 mover,
    # where K=J, so L = J*170). The template's L25/L33 formulas do the rest.
    final_price = extract_final_price_from_title(card_name)
    if final_price is None:
        total_hours = None
        result["warnings"].append(
            "Could not find a $ amount in card title. Hours cell will be blank — fill manually."
        )
    else:
        total_hours = round(final_price / 170.0, 4)

    # --- 3. Move date ---
    move_date_iso = find_move_to_list_date(card_id, INVOICE_BEP_COMPLETED_LIST, api_key, api_token)
    if move_date_iso:
        try:
            move_date = datetime.fromisoformat(move_date_iso.replace("Z", "+00:00"))
        except Exception:
            move_date = datetime.now()
            result["warnings"].append("Could not parse move-to-BEP-Completed date; using today.")
    else:
        move_date = datetime.now()
        result["warnings"].append(f"Card has no move-to-'{INVOICE_BEP_COMPLETED_LIST}' action; using today as move date.")

    move_date_str = _format_move_date(move_date)

    # --- 4. Open spreadsheet ---
    gc, creds = _get_gspread_client()
    if gc is None:
        result["error"] = (
            "Google service account not configured. "
            "Add GOOGLE_SERVICE_ACCOUNT_JSON to Railway env vars and share the template spreadsheet "
            "with the service account's email as Editor."
        )
        return result

    try:
        spreadsheet = gc.open_by_key(INVOICE_SPREADSHEET_ID)
    except Exception as e:
        result["error"] = f"Could not open invoice spreadsheet: {type(e).__name__}: {e}"
        return result

    # --- 5. Compute next invoice number and tab name ---
    invoice_num = get_next_invoice_number(spreadsheet)
    short_desc = clean_card_title_for_tab(card_name, bep_auth=bep_auth)
    tab_name = f"INV{invoice_num} {short_desc}".strip()
    # Google Sheets tab names max 100 chars
    tab_name = tab_name[:99]
    result["invoice_number"] = invoice_num
    result["tab_name"] = tab_name

    # --- 6. Duplicate template to the right of TEMPLATE ---
    template_tab_name = INVOICE_JOB_TYPES.get(job_type_label, INVOICE_TEMPLATE_TAB)
    try:
        template_ws = spreadsheet.worksheet(template_tab_name)
    except Exception as e:
        result["error"] = f"Template tab '{template_tab_name}' not found: {e}"
        return result

    # Find template's index, place new tab at index+1 (right side)
    all_ws = spreadsheet.worksheets()
    template_index = next((i for i, w in enumerate(all_ws) if w.id == template_ws.id), 0)
    try:
        new_ws = spreadsheet.duplicate_sheet(
            source_sheet_id=template_ws.id,
            insert_sheet_index=template_index + 1,
            new_sheet_name=tab_name,
        )
    except Exception as e:
        result["error"] = f"Failed to duplicate template: {type(e).__name__}: {e}"
        return result

    result["tab_gid"] = new_ws.id
    result["sheet_url"] = (
        f"https://docs.google.com/spreadsheets/d/{INVOICE_SPREADSHEET_ID}/edit#gid={new_ws.id}"
    )

    # --- 7. If more than 4 machines, insert extra blocks ---
    # Template has 4 pre-drawn blocks at rows 12, 16, 20, 24.
    # For N > 4, insert (N-4)*4 rows before row 24, then copy the structure of
    # block 1 (rows 12-15) into each new slot. The original block 4 (which has
    # the J25/K25/L25 formulas) shifts down; its formulas auto-update.
    n_machines = len(machines)
    extra_blocks = max(0, n_machines - 4)
    if extra_blocks > 0:
        try:
            requests_body = []
            # Insert blank rows before row 24 (0-indexed: row 23)
            requests_body.append({
                "insertDimension": {
                    "range": {
                        "sheetId": new_ws.id,
                        "dimension": "ROWS",
                        "startIndex": 23,
                        "endIndex": 23 + extra_blocks * 4,
                    },
                    "inheritFromBefore": False,
                }
            })
            # Copy block 1 (rows 12-15, 0-indexed 11-15) into each new slot
            for i in range(extra_blocks):
                dest_start = 23 + i * 4  # 0-indexed
                requests_body.append({
                    "copyPaste": {
                        "source": {
                            "sheetId": new_ws.id,
                            "startRowIndex": 11,
                            "endRowIndex": 15,
                            "startColumnIndex": 0,
                            "endColumnIndex": 12,
                        },
                        "destination": {
                            "sheetId": new_ws.id,
                            "startRowIndex": dest_start,
                            "endRowIndex": dest_start + 4,
                            "startColumnIndex": 0,
                            "endColumnIndex": 12,
                        },
                        "pasteType": "PASTE_NORMAL",
                    }
                })
            spreadsheet.batch_update({"requests": requests_body})
            # After insertion, re-fetch the worksheet to pick up updated indices
            new_ws = spreadsheet.worksheet(tab_name)
        except Exception as e:
            result["warnings"].append(
                f"Failed to expand template for {n_machines} machines: {type(e).__name__}: {e}. "
                f"Only the first 4 machines will be filled."
            )
            n_machines = 4
            machines = machines[:4]

    # --- 8. Fill cells ---
    try:
        header_updates = [
            {"range": f"{tab_name}!L7", "values": [[f"INVOICE # {invoice_num}"]]},
            {"range": f"{tab_name}!L5", "values": [[move_date_str]]},
            {"range": f"{tab_name}!F7", "values": [[bep_auth or ""]]},
            {"range": f"{tab_name}!C29", "values": [[note_text or ""]]},
        ]
        spreadsheet.values_batch_update({
            "valueInputOption": "USER_ENTERED",
            "data": header_updates,
        })
    except Exception as e:
        result["warnings"].append(f"Header fill failed: {type(e).__name__}: {e}")

    # How many blocks exist after possible row insertion. Template always has 4
    # pre-drawn blocks; if we inserted extra rows, there are now (4 + extra_blocks)
    # block slots, with the final one (originally block 4) containing the J/K/L
    # totals formulas at row (24 + extra_blocks*4).
    total_blocks = 4 + extra_blocks
    final_block_row = 24 + extra_blocks * 4  # row where the totals formulas live

    # Machine blocks — row of block i (1-indexed) = 12 + 4*(i-1)
    for i, machine in enumerate(machines, start=1):
        block_row = 12 + 4 * (i - 1)
        try:
            _fill_machine_block(new_ws, block_row, i, machine)
        except Exception as e:
            result["warnings"].append(f"Machine {i} fill failed: {type(e).__name__}: {e}")

    # Clear leftover template data from unused blocks (when N < total_blocks).
    # Don't touch rows in the final block's J/K/L columns — those hold formulas.
    for i in range(len(machines) + 1, total_blocks + 1):
        blank_row = 12 + 4 * (i - 1)
        blank_updates = [
            {"range": f"A{blank_row}", "values": [[""]]},
            {"range": f"B{blank_row}", "values": [[""]]},
            {"range": f"C{blank_row}", "values": [[""]]},
            {"range": f"H{blank_row}", "values": [[""]]},
            {"range": f"I{blank_row}", "values": [[""]]},
            {"range": f"B{blank_row + 1}", "values": [[""]]},
            {"range": f"C{blank_row + 1}", "values": [[""]]},
            {"range": f"B{blank_row + 2}", "values": [[""]]},
            {"range": f"C{blank_row + 2}", "values": [[""]]},
        ]
        try:
            new_ws.batch_update(blank_updates, value_input_option="USER_ENTERED")
        except Exception as e:
            result["warnings"].append(f"Clear unused block {i} failed: {type(e).__name__}: {e}")

    # Total hours — ALWAYS goes in J of the final block (where the L33/L25 formulas reference)
    hours_cell = f"J{final_block_row + 1}"
    if total_hours is not None:
        try:
            new_ws.update(hours_cell, [[total_hours]], value_input_option="USER_ENTERED")
        except Exception as e:
            result["warnings"].append(f"Hours fill failed: {type(e).__name__}: {e}")

    # --- 9. Read computed total from L33 (Balance Due) ---
    try:
        balance_val = new_ws.acell("L33", value_render_option="UNFORMATTED_VALUE").value
        if isinstance(balance_val, (int, float)):
            result["total_amount"] = float(balance_val)
        else:
            try:
                result["total_amount"] = float(str(balance_val).replace("$", "").replace(",", ""))
            except Exception:
                result["total_amount"] = None
    except Exception:
        result["total_amount"] = None

    # --- 10. Export tab as PDF ---
    try:
        pdf_bytes = export_sheet_tab_as_pdf(INVOICE_SPREADSHEET_ID, new_ws.id, creds)
        result["pdf_bytes"] = pdf_bytes
        result["pdf_filename"] = f"INV{invoice_num}.pdf"
    except Exception as e:
        result["warnings"].append(f"PDF export failed: {type(e).__name__}: {e}")

    # --- 11. Attach PDF to Trello card ---
    if result["pdf_bytes"]:
        try:
            attached = attach_pdf_to_card(
                card_id, result["pdf_bytes"], result["pdf_filename"], api_key, api_token
            )
            if not attached:
                result["warnings"].append("PDF attach to Trello card failed.")
        except Exception as e:
            result["warnings"].append(f"PDF attach error: {type(e).__name__}: {e}")

    # --- 12. Append row to Pending Payments ---
    try:
        pending_ws = spreadsheet.worksheet(INVOICE_PENDING_TAB)
        last_row = _col_last_data_row(pending_ws, col_index=1)
        append_row_index = last_row + 1
        today_str = datetime.now().strftime("%Y-%m-%d")
        row_values = [
            invoice_num,                                 # A Invoice
            today_str,                                   # B Date Sent
            result["total_amount"] if result["total_amount"] is not None else "",  # C Amount
            "Pending",                                   # D Status
            bep_auth or "",                              # E Auth
            "",                                          # F Item (blank per spec)
        ]
        pending_ws.update(
            f"A{append_row_index}:F{append_row_index}",
            [row_values],
            value_input_option="USER_ENTERED",
        )
    except Exception as e:
        result["warnings"].append(f"Pending Payments append failed: {type(e).__name__}: {e}")

    result["ok"] = True
    return result

def export_sheet_tab_as_pdf(spreadsheet_id, sheet_gid, creds):
    """
    Export ONE tab (by gid) of a Google Sheets file as PDF using the Drive export URL.
    Requires service account creds with Sheets + Drive scopes.
    """
    # Refresh the access token so the Authorization header is current
    from google.auth.transport.requests import Request
    if not creds.valid:
        creds.refresh(Request())
    token = creds.token
    url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export"
        f"?format=pdf&gid={sheet_gid}&portrait=true&size=letter&fitw=true&gridlines=false"
        f"&printtitle=false&sheetnames=false&pagenum=UNDEFINED&attachment=false"
    )
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"Sheets PDF export HTTP {r.status_code}: {r.text[:200]}")
    return r.content

# =============================================================================
# STREAMLIT UI
# =============================================================================

# Get Trello credentials once
trello_key = os.environ.get("TRELLO_API_KEY", "")
trello_token = os.environ.get("TRELLO_TOKEN", "")
trello_list = os.environ.get("TRELLO_LIST_ID", "699c9f9d6117bdcbb2d0e0aa")

# Sidebar - Page Navigation
with st.sidebar:
    st.title("🚚 BEP Tools")
    
    # Show logged in user and logout button
    st.caption(f"👤 Logged in as: **{st.session_state.get('username', 'unknown')}**")
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state["authenticated"] = False
        st.session_state["username"] = None
        st.rerun()
    
    st.divider()
    
    page = st.radio("Select Page:", ["📤 New Request", "📧 From Email", "📝 Generate Quote", "📄 Generate INV", "📊 Learning Data", "🗺️ Route Cache"], label_visibility="collapsed")
    
    st.divider()
    
    st.header("📋 Pricing Rules")
    st.markdown(f"""
    **Rate:** ${HOURLY_RATE}/hour
    **Job Time:** 30 min/machine
    **Buffer:** +20 min if >35 miles
    **No-buffer discount:** -$60 if ≤35 miles
    
    **Minimums:**
    - General: $220
    - Tucson: $850
    - Prison: $900+
    
    **Rounding:** Up to $25
    """)
    
    st.divider()
    
    with st.expander("🔧 Trello Settings", expanded=False):
        if trello_key and trello_token:
            st.success("✅ Trello credentials loaded")
        else:
            st.warning("⚠️ Set TRELLO_API_KEY and TRELLO_TOKEN in Railway variables")

# =============================================================================
# PAGE 2: FROM EMAIL
# =============================================================================
if page == "📧 From Email":
    st.title("📧 Process from Email")
    st.markdown("**Select a forwarded email → Auto-extract Excel & subject → Create card**")
    
    st.divider()
    
    if not GMAIL_APP_PASSWORD:
        st.error("⚠️ Gmail not configured. Add GMAIL_APP_PASSWORD to Railway variables.")
        st.markdown("""
        **Setup steps:**
        1. Go to [Google Account Security](https://myaccount.google.com/security)
        2. Enable 2-Factor Authentication
        3. Create an App Password for "Mail"
        4. Add to Railway: `GMAIL_APP_PASSWORD` = your app password
        """)
    else:
        with st.spinner("Connecting to Gmail..."):
            mail = connect_to_gmail()
        
        if mail:
            st.success("✅ Connected to Gmail")
            
            # Fetch recent emails with Excel
            with st.spinner("Fetching recent emails with Excel attachments..."):
                emails = get_recent_emails_with_excel(mail, limit=20)
            
            if emails:
                st.markdown(f"### Found {len(emails)} email(s) with Excel attachments")
                
                for i, em in enumerate(emails):
                    with st.expander(f"📩 {em['subject'][:60]}...", expanded=(i==0)):
                        st.markdown(f"**From:** {em['from']}")
                        st.markdown(f"**Date:** {em['date']}")
                        st.markdown(f"**Subject:** {em['subject']}")
                        
                        if st.button(f"📥 Process this email", key=f"process_{em['id']}", use_container_width=True):
                            # Extract ALL Excel attachments
                            all_excels = get_all_excels_from_email(em['message'])
                            
                            if all_excels:
                                # Identify each Excel type (MR vs WO)
                                mr_file = None
                                wo_files = []
                                
                                for excel in all_excels:
                                    file_type = identify_excel_type(excel['filename'], excel['data'])
                                    if file_type == 'MR' and mr_file is None:
                                        mr_file = excel
                                    elif file_type == 'WO':
                                        wo_files.append(excel)
                                    elif mr_file is None:
                                        # If type unknown and no MR yet, assume first is MR
                                        mr_file = excel
                                    else:
                                        # Additional unknown files treated as WO
                                        wo_files.append(excel)
                                
                                # Store in session
                                if mr_file:
                                    st.session_state['email_excel'] = mr_file['data']
                                    st.session_state['email_excel_name'] = mr_file['filename']
                                st.session_state['email_wo_files'] = wo_files
                                st.session_state['email_subject'] = em['subject']
                                st.session_state['has_workorder'] = len(wo_files) > 0
                                
                                # Show what was found
                                if mr_file:
                                    st.success(f"✅ Move Request: {mr_file['filename']}")
                                if wo_files:
                                    for wo in wo_files:
                                        st.info(f"📋 Work Order: {wo['filename']}")
                
                # If email selected, show processing UI
                if 'email_excel' in st.session_state:
                    st.divider()
                    st.markdown("### 📋 Process Selected Email")
                    
                    st.info(f"**Subject:** {st.session_state.get('email_subject', '')}")
                    st.info(f"**Move Request:** {st.session_state.get('email_excel_name', '')}")
                    
                    # Show Work Order info if present
                    has_workorder = st.session_state.get('has_workorder', False)
                    wo_files = st.session_state.get('email_wo_files', [])
                    if has_workorder and wo_files:
                        st.warning(f"📋 **HAS WORK ORDER:** {', '.join([w['filename'] for w in wo_files])}")
                    
                    # Parse the Excel
                    excel_bytes = st.session_state['email_excel']
                    with st.spinner("Parsing Excel..."):
                        uploaded_file = BytesIO(excel_bytes)
                        data = parse_bep_excel_v2(uploaded_file)
                    
                    if data['success']:
                        num_machines = len(data.get('machines', []))
                        if num_machines > 0:
                            st.success(f"✅ Extracted {num_machines} machine(s)")
                        else:
                            st.warning("⚠️ No machines found in Excel. Check if this is a Move Request file with pickup/delivery addresses.")
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Editable fields
                            card_title = st.text_input("Card Title", value=st.session_state.get('email_subject', ''))
                            requester = st.text_input("Requester", value=data.get('requester') or '')
                            mr_number = st.text_input("MR Number", value=data.get('mr_number') or '')
                            
                            st.markdown("**Machines:**")
                            for m in data.get('machines', []):
                                st.caption(f"• {m.get('type', 'Machine')}: {m.get('pickup', '?')} → {m.get('delivery', '?')}")
                        
                        with col2:
                            # Get unique addresses
                            machines_list = data.get('machines', [])
                            unique_pickups = clean_and_dedupe_addresses([m["pickup"] for m in machines_list if m.get("pickup")])
                            unique_deliveries = clean_and_dedupe_addresses([m["delivery"] for m in machines_list if m.get("delivery")])
                            
                            if not unique_pickups and not unique_deliveries:
                                st.error("❌ No pickup or delivery addresses found. Cannot calculate quote.")
                                st.info("This Excel may be a Work Order only (no Move Request data).")
                            elif st.button("🧮 Calculate Quote", type="primary", use_container_width=True):
                                with st.spinner("Calculating route..."):
                                    route_data = calculate_optimal_route(machines_list)
                                    quote = calculate_quote(route_data, len(machines_list), unique_pickups, unique_deliveries)
                                    
                                    st.session_state['email_quote'] = quote
                                    st.session_state['email_route'] = route_data
                                    st.session_state['email_data'] = data
                            
                            if 'email_quote' in st.session_state:
                                quote = st.session_state['email_quote']
                                st.success(f"### 💵 Quote: ${quote['final_price']:,}")
                                st.caption(f"Drive: {quote['drive_time']}min + Job: {quote['job_time']}min + Buffer: {quote['buffer_time']}min")
                                
                                # Show smart adjustments if any
                                if quote.get('smart_adjustment', 0) != 0:
                                    reasons = quote.get('adjustment_reasons', [])
                                    st.info(f"📊 **Smart Adjustment:** +${quote['smart_adjustment']} ({', '.join(reasons)})")
                                
                                # Build full title (with HAS WORKORDER flag if applicable)
                                has_workorder = st.session_state.get('has_workorder', False)
                                wo_flag = " - HAS WORKORDER" if has_workorder else ""
                                if requester:
                                    full_title = f"{requester}{wo_flag} - {card_title} - ${quote['final_price']}"
                                else:
                                    full_title = f"{card_title}{wo_flag} - ${quote['final_price']}"
                                st.text_input("Final Card Title", value=full_title, key="final_title")
                                
                                if trello_key and trello_token:
                                    if st.button("📋 Create Trello Card + Attach Files", use_container_width=True, type="primary"):
                                        with st.spinner("Creating card..."):
                                            full_data = {
                                                "requester": requester,
                                                "mr_number": mr_number,
                                                "machines": data.get('machines', []),
                                                "unique_pickups": unique_pickups,
                                                "unique_deliveries": unique_deliveries,
                                                "num_machines": len(data.get('machines', [])),
                                                "other_notes": data.get('other_notes'),
                                                "route": (st.session_state.get('email_route') or {}).get("route", []),
                                                **quote
                                            }
                                            
                                            # Override title
                                            full_data['card_title'] = st.session_state.get('final_title', full_title)
                                            
                                            card = create_trello_card(full_data, trello_key, trello_token, trello_list)
                                            
                                            if card:
                                                card_id = card.get('id')
                                                st.success(f"✅ Card created!")
                                                
                                                # ATTACHMENT ORDER: 
                                                # 1. CAPTURE PDF (easiest for Ryan on mobile)
                                                # 2. Excel file(s)
                                                # 3. Work Order PDF (if exists)
                                                
                                                # 1. Generate and attach CAPTURE PDF FIRST
                                                with st.spinner("Generating CAPTURE PDF..."):
                                                    pdf = convert_excel_to_pdf(excel_bytes, st.session_state.get('email_excel_name', 'request.xlsx'))
                                                    if pdf:
                                                        pdf = remove_pdf_pages(pdf, [1])
                                                        pdf_name = f"CAPTURE_{mr_number or 'BEP'}_{datetime.now().strftime('%Y%m%d')}.pdf"
                                                        attach_pdf_to_card(card_id, pdf, pdf_name, trello_key, trello_token)
                                                        st.success("✅ CAPTURE PDF attached")
                                                
                                                # 2. Attach Move Request Excel
                                                with st.spinner("Attaching Excel..."):
                                                    attach_excel_to_card(card_id, excel_bytes, st.session_state.get('email_excel_name', 'request.xlsx'), trello_key, trello_token)
                                                
                                                # 3. Attach Work Order files (if any)
                                                wo_files = st.session_state.get('email_wo_files', [])
                                                if wo_files:
                                                    for wo in wo_files:
                                                        with st.spinner(f"Processing Work Order: {wo['filename']}..."):
                                                            # Attach WO Excel
                                                            attach_excel_to_card(card_id, wo['data'], wo['filename'], trello_key, trello_token)
                                                            
                                                            # Convert WO to PDF and attach
                                                            wo_pdf = convert_workorder_to_pdf(wo['data'], wo['filename'])
                                                            if wo_pdf:
                                                                wo_pdf_name = f"WORKORDER_{os.path.splitext(wo['filename'])[0]}.pdf"
                                                                attach_pdf_to_card(card_id, wo_pdf, wo_pdf_name, trello_key, trello_token)
                                                                st.success(f"✅ Work Order PDF attached: {wo_pdf_name}")
                                                
                                                st.success("✅ All files attached!")
                                                st.markdown(f"[Open Card]({card.get('shortUrl')})")
                                                
                                                # Clear session
                                                for key in ['email_excel', 'email_excel_name', 'email_subject', 'email_quote', 'email_route', 'email_data', 'email_wo_files', 'has_workorder']:
                                                    if key in st.session_state:
                                                        del st.session_state[key]
                                            else:
                                                st.error("Failed to create card")
                    else:
                        st.error(f"Failed to parse Excel: {data.get('error')}")
            else:
                st.info("No recent emails with Excel attachments found.")
            
            mail.logout()
        else:
            st.error("Could not connect to Gmail")

# =============================================================================
# PAGE 3: GENERATE QUOTE PDF
# =============================================================================
elif page == "📝 Generate Quote":
    st.title("📝 Generate Quote PDF")
    st.markdown("**After Ryan approves → Generate QUOTE PDF with filled worksheet**")
    
    st.divider()
    
    # Input: Card URL or ID
    card_input = st.text_input(
        "Trello Card URL or ID",
        placeholder="https://trello.com/c/ABC123 or card ID",
        help="Paste the Trello card URL or just the card ID"
    )
    
    # Extract card ID from URL
    card_id = None
    if card_input:
        # Handle full URL
        match = re.search(r'/c/([a-zA-Z0-9]+)', card_input)
        if match:
            card_id = match.group(1)
        else:
            # Assume it's just the ID
            card_id = card_input.strip()
    
    if card_id and trello_key and trello_token:
        # Fetch card info
        card_info = get_card_info(card_id, trello_key, trello_token)
        
        if card_info:
            st.success(f"✅ Found card: **{card_info.get('name')}**")
            
            card_title = card_info.get('name', '')
            
            # FIRST: Check if title has adjustment pattern like "$350 change to $400"
            title_original, title_final = extract_price_adjustment_from_title(card_title)
            
            # SECOND: Try to extract from description marker or quote section
            desc_original = extract_original_quote_from_desc(card_info.get('desc', ''))
            
            # THIRD: Extract final price from title (just the last $XXX)
            auto_quote = extract_quote_from_title(card_title)
            
            # FOURTH: Check comments for price adjustments
            comments = get_card_comments(card_id, trello_key, trello_token)
            comment_price = extract_price_adjustment_from_comments(comments)
            comment_text = " | ".join(comments) if comments else ""
            
            # Determine which original/final to use
            if title_original and title_final:
                # Title has explicit adjustment pattern - use that
                original_quote = title_original
                final_quote = title_final
                auto_quote = title_final  # For the input field
                st.info(f"📊 **Title shows adjustment:** ${title_original} → ${title_final}")
            elif comment_price and auto_quote:
                # Comment has price adjustment
                original_quote = auto_quote
                final_quote = comment_price
                auto_quote = comment_price  # For the input field
                st.info(f"📊 **Comment shows adjustment:** ${original_quote} → ${comment_price}")
            elif desc_original:
                # Use description marker vs title price
                original_quote = desc_original
                final_quote = auto_quote
            else:
                # No adjustment info available
                original_quote = None
                final_quote = auto_quote
            
            # Show comments if any
            if comments:
                with st.expander(f"💬 Card Comments ({len(comments)})"):
                    for c in comments[:5]:  # Show first 5 comments
                        st.caption(c[:200] + "..." if len(c) > 200 else c)
            
            locations = extract_locations_from_desc(card_info.get('desc', ''))
            
            # Show learning feedback if there's a difference
            if original_quote and final_quote and original_quote != final_quote:
                diff = final_quote - original_quote
                diff_pct = round((diff / original_quote) * 100, 1)
                if diff > 0:
                    st.warning(f"📊 **Price Adjustment Detected:** Original ${original_quote} → Final ${final_quote} (**+${diff}**, +{diff_pct}%)")
                else:
                    st.info(f"📊 **Price Adjustment Detected:** Original ${original_quote} → Final ${final_quote} (**${diff}**, {diff_pct}%)")
            elif original_quote and final_quote and original_quote == final_quote:
                st.success(f"✅ **Price Match:** Calculated ${original_quote} = Final ${final_quote}")
            elif not original_quote:
                st.info("ℹ️ No original quote found - this card was created before the learning system")
            
            col1, col2 = st.columns(2)
            with col1:
                quote_amount = st.number_input(
                    "Quote Amount ($)",
                    min_value=100,
                    max_value=10000,
                    value=auto_quote or 300,
                    step=25
                )
            with col2:
                signature = st.text_input("Signature", value="Ryan Kearl")
            
            # Show calculated hours
            hours = round(quote_amount / HOURLY_RATE, 2)
            st.info(f"**Calculated Hours:** {hours} hrs (${quote_amount} ÷ ${HOURLY_RATE})")
            
            # Fetch attachments
            attachments = get_card_attachments(card_id, trello_key, trello_token)
            excel_attachments = [a for a in attachments if a.get('name', '').endswith(('.xlsx', '.xls'))]
            
            if excel_attachments:
                st.success(f"✅ Found Excel file: **{excel_attachments[0].get('name')}**")
                
                if st.button("🧮 Generate QUOTE PDF", type="primary", use_container_width=True):
                    with st.spinner("Downloading Excel..."):
                        excel_url = excel_attachments[0].get('url')
                        excel_bytes = download_attachment(excel_url, trello_key, trello_token)
                    
                    if excel_bytes:
                        with st.spinner("Filling worksheet & generating PDF..."):
                            pdf_bytes, calc_hours = fill_worksheet_and_generate_pdf(
                                excel_bytes, quote_amount, signature
                            )
                        
                        if pdf_bytes:
                            # Store in session state for persistence
                            st.session_state['quote_pdf'] = pdf_bytes
                            st.session_state['quote_pdf_card_id'] = card_id
                            st.session_state['quote_pdf_card_url'] = card_info.get('shortUrl')
                            st.session_state['quote_pdf_hours'] = calc_hours
                            st.success(f"✅ QUOTE PDF generated! ({calc_hours} hours)")
                        else:
                            st.error("Failed to generate PDF")
                    else:
                        st.error("Failed to download Excel file")
                
                # Show PDF options if generated
                if 'quote_pdf' in st.session_state and st.session_state.get('quote_pdf_card_id') == card_id:
                    st.success(f"✅ QUOTE PDF ready! ({st.session_state.get('quote_pdf_hours', '')} hours)")

                    safe_title = re.sub(r'[\\/:*?"<>|]+', '_', card_info.get('name', '') or card_id).strip().strip('.')
                    safe_title = f"QUOTE {safe_title[:150]}" if safe_title else f"QUOTE {card_id}"

                    col_a, col_b = st.columns(2)

                    with col_a:
                        st.download_button(
                            "⬇️ Download QUOTE PDF",
                            data=st.session_state['quote_pdf'],
                            file_name=f"{safe_title}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )

                    with col_b:
                        if st.button("📎 Attach to Trello Card", use_container_width=True):
                            with st.spinner("Attaching..."):
                                filename = f"{safe_title}.pdf"
                                if attach_pdf_to_card(card_id, st.session_state['quote_pdf'], filename, trello_key, trello_token):
                                    st.success(f"✅ QUOTE PDF attached!")
                                    st.markdown(f"[Open Card]({st.session_state.get('quote_pdf_card_url')})")
                                    
                                    # Log feedback for learning system
                                    if original_quote:
                                        # comments and comment_text already fetched above
                                        diff = log_quote_feedback(
                                            original_quote=original_quote,
                                            final_price=quote_amount,
                                            card_name=card_info.get('name', ''),
                                            locations=locations,
                                            comments=comment_text
                                        )
                                        if diff != 0:
                                            st.info(f"📊 Logged price adjustment: ${diff:+d} (learning system updated)")
                                            
                                            # Request AI analysis from Grant
                                            if post_analysis_request(
                                                card_id=card_id,
                                                original_quote=original_quote,
                                                final_price=quote_amount,
                                                ryan_comment=comment_text,
                                                api_key=trello_key,
                                                api_token=trello_token
                                            ):
                                                st.success("🤖 AI analysis requested - Grant will review within 6 hours")
                                            else:
                                                st.warning("⚠️ Could not request AI analysis")
                                        else:
                                            st.info("📊 Logged: Quote matched (no adjustment)")
                                    
                                    # Clear the session state
                                    del st.session_state['quote_pdf']
                                else:
                                    st.error("Failed to attach PDF")
            else:
                st.warning("⚠️ No Excel file found on this card. Upload one first on Page 1.")
        else:
            st.error("❌ Card not found. Check the URL/ID.")
    elif card_id and not (trello_key and trello_token):
        st.error("⚠️ Trello credentials not configured")

# =============================================================================
# PAGE: GENERATE INVOICE
# =============================================================================
elif page == "📄 Generate INV":
    st.title("📄 Generate INV")
    st.markdown("**Create a Google Sheets invoice from a Trello card, export PDF, attach back to card**")

    # Service account status
    sa_set = bool(os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip())
    if not sa_set:
        st.error(
            "⚠️ Google service account not configured. "
            "Add `GOOGLE_SERVICE_ACCOUNT_JSON` to Railway env vars and share the template "
            f"spreadsheet with the service account's email as Editor. "
            f"Template: https://docs.google.com/spreadsheets/d/{INVOICE_SPREADSHEET_ID}/"
        )
    elif not (trello_key and trello_token):
        st.error("⚠️ Trello credentials missing. Set TRELLO_API_KEY and TRELLO_TOKEN.")
    else:
        st.success("✅ Service account and Trello credentials detected.")

    st.divider()

    card_input = st.text_input(
        "Trello Card URL or ID",
        placeholder="https://trello.com/c/ABC123 or card ID",
    )
    job_type = st.selectbox("Job Type", list(INVOICE_JOB_TYPES.keys()))
    note_text = st.text_area("Note (optional — goes into C29)", value="", height=80)

    # Resolve card ID
    card_id_inv = None
    if card_input:
        m = re.search(r'/c/([a-zA-Z0-9]+)', card_input)
        card_id_inv = m.group(1) if m else card_input.strip()

    if card_id_inv and trello_key and trello_token and sa_set:
        if st.button("🧾 Generate Invoice", type="primary", use_container_width=True):
            with st.spinner("Generating invoice…"):
                result = generate_invoice_from_card(
                    card_id_inv, job_type, note_text, trello_key, trello_token
                )

            if result.get("error"):
                st.error(f"❌ {result['error']}")
                if result.get("raw_desc"):
                    with st.expander("🔍 Raw card description (debug)", expanded=False):
                        st.code(result["raw_desc"], language="text")
            else:
                st.success(f"✅ Invoice **INV{result['invoice_number']}** generated!")
                col_a, col_b = st.columns(2)
                with col_a:
                    if result.get("sheet_url"):
                        st.markdown(f"[📊 Open new sheet tab]({result['sheet_url']})")
                    if result.get("total_amount") is not None:
                        st.metric("Balance Due", f"${result['total_amount']:,.2f}")
                with col_b:
                    if result.get("pdf_bytes"):
                        st.download_button(
                            "⬇️ Download Invoice PDF",
                            data=result["pdf_bytes"],
                            file_name=result["pdf_filename"],
                            mime="application/pdf",
                            use_container_width=True,
                        )
                if result.get("warnings"):
                    st.warning("Completed with warnings:")
                    for w in result["warnings"]:
                        st.caption(f"• {w}")

# =============================================================================
# PAGE 4: LEARNING DATA
# =============================================================================
elif page == "📊 Learning Data":
    st.title("📊 Learning Data")
    st.markdown("**View and manage the pricing learning system**")
    
    st.divider()
    
    # Load learning data
    data = load_learning_data()
    
    # Summary stats
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Quotes", data.get("total_quotes", 0))
    with col2:
        st.metric("Adjustments", data.get("total_adjustments", 0))
    with col3:
        pct = round((data.get("total_adjustments", 0) / max(data.get("total_quotes", 1), 1)) * 100, 1)
        st.metric("Adjustment Rate", f"{pct}%")
    with col4:
        st.metric("Avg Adjustment", f"${data.get('avg_adjustment', 0):+.0f}")
    
    st.divider()
    
    # Location stats
    st.subheader("📍 Location-Based Adjustments")
    loc_stats = data.get("location_stats", {})
    if loc_stats:
        loc_df = pd.DataFrame([
            {"Location Type": k, "Count": v["count"], "Avg Adjustment": f"${v['avg_diff']:+d}"}
            for k, v in loc_stats.items()
        ])
        st.dataframe(loc_df, use_container_width=True, hide_index=True)
    else:
        st.info("No location-specific data yet. Process some quotes to build learning data.")
    
    st.divider()
    
    # Recent quotes
    st.subheader("📜 Recent Quote History")
    quotes = data.get("quotes", [])
    if quotes:
        # Show last 20 quotes
        recent = quotes[-20:][::-1]  # Reverse to show newest first
        
        for q in recent:
            adj = q.get("adjustment", 0)
            if adj > 0:
                icon = "🔺"
                color = "orange"
            elif adj < 0:
                icon = "🔻"
                color = "blue"
            else:
                icon = "✅"
                color = "green"
            
            with st.expander(f"{icon} {q.get('card_name', 'Unknown')[:50]}... | ${q.get('original_quote', 0)} → ${q.get('final_price', 0)} ({adj:+d})"):
                st.write(f"**Original Quote:** ${q.get('original_quote', 0)}")
                st.write(f"**Final Price:** ${q.get('final_price', 0)}")
                st.write(f"**Adjustment:** ${adj:+d} ({q.get('adjustment_pct', 0):+.1f}%)")
                st.write(f"**Timestamp:** {q.get('timestamp', 'N/A')}")
                if q.get('comments'):
                    st.write(f"**Comments:** {q.get('comments')}")
                if q.get('locations'):
                    st.write(f"**Locations:** {', '.join(q.get('locations', []))[:100]}...")
    else:
        st.info("No quotes logged yet. Generate some QUOTE PDFs to start learning.")
    
    st.divider()
    
    # Raw JSON view
    with st.expander("🔧 Raw Learning Data (JSON)"):
        st.json(data)
    
    # Download link
    st.download_button(
        "⬇️ Download Learning Data (JSON)",
        data=json.dumps(data, indent=2),
        file_name=f"learning_data_{datetime.now().strftime('%Y%m%d')}.json",
        mime="application/json"
    )
    
    # Reset option (with confirmation)
    st.divider()
    with st.expander("⚠️ Danger Zone"):
        st.warning("This will delete all learning data and start fresh.")
        if st.button("🗑️ Reset Learning Data", type="secondary"):
            if st.session_state.get('confirm_reset'):
                # Actually reset
                save_learning_data({
                    "version": "1.0",
                    "created": datetime.now().isoformat(),
                    "quotes": [],
                    "location_stats": {},
                    "total_quotes": 0,
                    "total_adjustments": 0,
                    "avg_adjustment": 0
                })
                st.success("Learning data reset!")
                st.session_state['confirm_reset'] = False
                st.rerun()
            else:
                st.session_state['confirm_reset'] = True
                st.warning("Click again to confirm reset")

# =============================================================================
# PAGE 1: NEW REQUEST (Original functionality)
# =============================================================================
elif page == "📤 New Request":
    st.title("📤 New Request")
    st.markdown("**Upload Excel → Calculate quote → Create Trello card**")
    
    # Main content
    st.header("Upload BEP Move Request")
    
    uploaded_file = st.file_uploader(
        "Drop your BEP Excel file here",
        type=["xlsx", "xls"],
        help="Upload the Move Request Excel file"
    )
    
    if uploaded_file:
        # Store Excel bytes for later attachment
        excel_bytes = uploaded_file.getvalue()
        excel_name = uploaded_file.name
        
        with st.spinner("Parsing Excel file..."):
            data = parse_bep_excel_v2(uploaded_file)
        
        if data["success"]:
            st.success(f"✅ Extracted {len(data['machines'])} machine(s)")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📋 Extracted Data")
                
                requester = st.text_input("Requester", value=data.get("requester") or "")
                mr_number = st.text_input("MR Number", value=data.get("mr_number") or "")
                move_date = st.text_input("Move Date", value=data.get("move_date") or "")
                
                st.markdown("### 🚛 Machines")
                
                # Editable machine list
                machines = data.get("machines", [])
                edited_machines = []
                
                for i, m in enumerate(machines):
                    with st.expander(f"Machine {m.get('number', i+1)}: {m.get('type', 'Vending')}", expanded=True):
                        pickup = st.text_input(f"Pickup {i+1}", value=m.get("pickup") or "", key=f"pickup_{i}")
                        delivery = st.text_input(f"Delivery {i+1}", value=m.get("delivery") or "", key=f"delivery_{i}")
                        mtype = st.text_input(f"Type {i+1}", value=m.get("type") or "Vending", key=f"type_{i}")
                        
                        edited_machines.append({
                            "number": i + 1,
                            "pickup": pickup,
                            "pickup_address": normalize_address(pickup),
                            "delivery": delivery,
                            "delivery_address": normalize_address(delivery),
                            "type": mtype
                        })
                
                # Add machine button
                if st.button("➕ Add Machine"):
                    edited_machines.append({
                        "number": len(edited_machines) + 1,
                        "pickup": "",
                        "pickup_address": "",
                        "delivery": "",
                        "delivery_address": "",
                        "type": "Vending"
                    })
                    st.rerun()
                
                other_notes = st.text_area("Other Notes", value=data.get("other_notes") or "", height=100)
            
            with col2:
                st.subheader("🗺️ Route & Quote")
                
                # Get unique addresses for routing
                unique_pickups = clean_and_dedupe_addresses([m["pickup"] for m in edited_machines if m.get("pickup")])
                unique_deliveries = clean_and_dedupe_addresses([m["delivery"] for m in edited_machines if m.get("delivery")])

                # True unique-stop count: a location that is both a pickup and a
                # delivery counts ONCE because the optimal router visits it once.
                all_stop_keys = {_dedupe_key(a) for a in unique_pickups + unique_deliveries if a}
                all_stop_keys.discard("")
                unique_stops = len(all_stop_keys)

                st.info(f"""
                **Route:** Gilbert, AZ 85295 → {unique_stops} unique stop(s) → HQ

                **Machines:** {len(edited_machines)}  ({len(unique_pickups)} distinct pickup addr, {len(unique_deliveries)} distinct delivery addr)
                """)
                
                if st.button("🧮 CALCULATE ROUTE & QUOTE", type="primary", use_container_width=True):
                    if unique_pickups or unique_deliveries:
                        with st.spinner("Calculating route via Google Maps..."):
                            route_data = calculate_optimal_route(edited_machines)
                            quote = calculate_quote(route_data, len(edited_machines), unique_pickups, unique_deliveries)
                            
                            st.session_state['route_data'] = route_data
                            st.session_state['quote'] = quote
                            st.session_state['excel_bytes'] = excel_bytes
                            st.session_state['excel_name'] = excel_name
                            st.session_state['full_data'] = {
                                "requester": requester,
                                "mr_number": mr_number,
                                "move_date": move_date,
                                "machines": edited_machines,
                                "unique_pickups": unique_pickups,
                                "unique_deliveries": unique_deliveries,
                                "num_machines": len(edited_machines),
                                "other_notes": other_notes,
                                "route": route_data.get("route", []),
                                **quote
                            }
                    else:
                        st.error("Need at least one pickup or delivery address")
                
                # Show results
                if 'quote' in st.session_state:
                    quote = st.session_state['quote']
                    route = st.session_state['route_data']
                    
                    st.success(f"### 💵 Quote: ${quote['final_price']:,}")
                    
                    # Show smart adjustments if any
                    if quote.get('smart_adjustment', 0) != 0:
                        reasons = quote.get('adjustment_reasons', [])
                        st.info(f"📊 **Smart Adjustment:** +${quote['smart_adjustment']} ({', '.join(reasons)})")
                    
                    # Route details
                    st.markdown("**Route Legs:**")
                    for leg in route['legs']:
                        est = " ⚠️" if leg.get('estimated') else ""
                        st.caption(f"• {leg['from'][:30]}... → {leg['to'][:30]}...: {leg['distance_miles']} mi, {leg['duration_minutes']} min{est}")
                    
                    # Indicators
                    col_a, col_b = st.columns(2)
                    with col_a:
                        if quote['is_tucson']:
                            st.warning("🌵 Tucson - $850 min")
                        if quote['is_prison']:
                            st.warning("🏛️ Prison - $900 min")
                    with col_b:
                        if quote['buffer_time'] > 0:
                            st.info(f"📍 Buffer: +{quote['buffer_time']} min (>{BUFFER_THRESHOLD_MILES} mi)")
                        else:
                            st.info(f"💰 No-buffer discount: -$60")
                    
                    # Breakdown
                    st.markdown(f"""
                    | Component | Value |
                    |-----------|-------|
                    | Drive Time | {quote['drive_time']} min |
                    | Job Time | {quote['job_time']} min |
                    | Buffer | {quote['buffer_time']} min |
                    | Max Leg Distance | {quote['max_distance_miles']} mi |
                    | **Total** | **{quote['total_hours']} hrs** |
                    """)
                    
                    st.caption(f"Formula: {quote['formula']}")
                    
                    st.divider()
                    
                    # Downloads
                    full_data = st.session_state.get('full_data', {})
                    
                    col_dl1, col_dl2 = st.columns(2)
                    
                    with col_dl1:
                        if st.button("📋 Convert Excel to PDF", use_container_width=True):
                            with st.spinner("Converting..."):
                                pdf = convert_excel_to_pdf(st.session_state.get('excel_bytes'), st.session_state.get('excel_name', 'request.xlsx'))
                                if pdf:
                                    pdf = remove_pdf_pages(pdf, [1])
                                    st.session_state['request_pdf'] = pdf
                                    st.success("✅ PDF ready!")
                    
                    with col_dl2:
                        if 'request_pdf' in st.session_state:
                            st.download_button(
                                "⬇️ Download CAPTURE PDF",
                                data=st.session_state['request_pdf'],
                                file_name=f"CAPTURE_{mr_number or 'BEP'}_{datetime.now().strftime('%Y%m%d')}.pdf",
                                mime="application/pdf",
                                use_container_width=True
                            )
                    
                    # Trello
                    if trello_key and trello_token:
                        if st.button("📋 Create Trello Card + Attach Files", use_container_width=True):
                            with st.spinner("Creating card..."):
                                card = create_trello_card(full_data, trello_key, trello_token, trello_list)
                                if card:
                                    card_id = card.get('id')
                                    st.success(f"✅ Card created: {card.get('shortUrl')}")
                                    
                                    # Attach Excel file (needed for Quote generation later)
                                    with st.spinner("Attaching Excel..."):
                                        if attach_excel_to_card(card_id, st.session_state.get('excel_bytes'), st.session_state.get('excel_name', 'request.xlsx'), trello_key, trello_token):
                                            st.success("✅ Excel attached!")
                                        else:
                                            st.warning("⚠️ Excel attachment failed")
                                    
                                    # Attach PDF if available
                                    if 'request_pdf' in st.session_state:
                                        with st.spinner("Attaching CAPTURE PDF..."):
                                            pdf_filename = f"CAPTURE_{mr_number or 'BEP'}_{datetime.now().strftime('%Y%m%d')}.pdf"
                                            if attach_pdf_to_card(card_id, st.session_state['request_pdf'], pdf_filename, trello_key, trello_token):
                                                st.success("✅ CAPTURE PDF attached!")
                                            else:
                                                st.warning("⚠️ PDF attachment failed")
                                    else:
                                        st.info("💡 Convert Excel to PDF first for CAPTURE attachment")
                                    
                                    st.markdown(f"**Next:** After Ryan approves, go to **📝 Generate Quote** page to create QUOTE PDF")
                                else:
                                    st.error("Failed to create Trello card")
            
            # Raw data viewer
            with st.expander("🔍 Raw Excel Data"):
                st.text("\n".join(data.get("raw_data", [])[:60]))
        
        else:
            st.error(f"❌ Error: {data.get('error')}")
    
    else:
        st.info("👆 Upload a BEP Move Request Excel file to get started")

elif page == "🗺️ Route Cache":
    st.title("🗺️ Route Cache")
    st.caption("Cached Google Maps route legs (stored in Supabase).")

    if not supabase:
        st.error("Supabase is not configured. Set SUPABASE_URL and SUPABASE_KEY in the environment.")
    else:
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button("🔄 Refresh"):
                st.rerun()
        with col3:
            search = st.text_input("Filter by origin/destination (optional)", "", label_visibility="collapsed", placeholder="Filter by origin or destination…")

        try:
            result = supabase.table("route_cache").select("*").order("created_at", desc=True).limit(1000).execute()
            rows = result.data or []

            if search:
                s = search.lower()
                rows = [r for r in rows if s in (r.get("origin") or "").lower() or s in (r.get("destination") or "").lower()]

            st.metric("Cached routes", len(rows))

            if rows:
                df = pd.DataFrame(rows)
                # Order columns nicely
                preferred = ["origin", "destination", "distance_miles", "duration_minutes", "created_at", "cache_key"]
                cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
                st.dataframe(df[cols], use_container_width=True, hide_index=True)

                with col2:
                    if st.button("🗑️ Clear cache", type="secondary"):
                        if st.session_state.get("_confirm_clear_cache"):
                            try:
                                supabase.table("route_cache").delete().neq("cache_key", "").execute()
                                st.session_state["_confirm_clear_cache"] = False
                                st.success("Cache cleared.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Clear failed: {e}")
                        else:
                            st.session_state["_confirm_clear_cache"] = True
                            st.warning("Click 'Clear cache' again to confirm.")
            else:
                st.info("No cached routes yet. Generate a quote to populate the cache.")
        except Exception as e:
            st.error(f"Could not load route cache: {type(e).__name__}: {e}")

# Footer
st.divider()
st.caption("BEP Pricing Calculator v4.0 | Two-Step Workflow | Tool Box & Safe Moving")
